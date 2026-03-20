"""
verify_wines.py
===============
Run this script with your XLSX file path to see:
  - Each wine name as parsed
  - The exact JR website URL to verify it manually

Usage:
  python verify_wines.py "C:/path/to/your/file.xlsx"
"""
import sys
import re
import urllib.parse

try:
    from openpyxl import load_workbook
except ImportError:
    print("Run:  pip install openpyxl")
    sys.exit(1)


def parse_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row
    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        row_low = [str(c or "").strip().lower().replace(" ", "_") for c in row]
        if any("product_name" in c or "lwin" in c for c in row_low):
            hdr_idx = i
            break

    raw_headers = [str(c or "").strip().lower().replace(" ", "_") for c in rows[hdr_idx]]

    def col(name_part):
        for i, h in enumerate(raw_headers):
            if name_part in h:
                return i
        return None

    name_col    = col("product_name") or col("name")
    vintage_col = col("vintage")
    lwin_col    = col("lwin")

    wines = []
    for row in rows[hdr_idx + 2:]:          # skip header + example row
        def cell(c):
            if c is None or c >= len(row) or row[c] is None:
                return ""
            return str(row[c]).strip()

        name = cell(name_col)
        if not name or "optional" in name.lower():
            continue

        vintage = cell(vintage_col)
        lwin    = cell(lwin_col)

        # Clean name (same as xlsx_service)
        name = re.sub(r"\s*\(.*?\)\s*$", "", name).strip()
        name = re.sub(r",?\s*\b(19|20)\d{2}\b\s*$", "", name).strip()

        # Derive vintage from LWIN if missing
        if not vintage and len(lwin) >= 11:
            lv = lwin[7:11]
            if lv.isdigit() and int(lv) > 1900:
                vintage = lv

        wines.append({"name": name, "vintage": vintage, "lwin": lwin})

    return wines


def jr_search_url(name, vintage=""):
    """Generate JancisRobinson.com tastings search URL for a wine."""
    # Use the most specific part of the name for search
    parts = [p.strip() for p in name.split(",") if p.strip()]

    if len(parts) >= 3:
        # 3-part: "Producer, Appellation, Cru"  → search "Producer Cru vintage"
        query = f"{parts[0]} {parts[2]}"
    elif len(parts) == 2:
        query = f"{parts[0]} {parts[1]}"
    else:
        query = name

    if vintage and vintage.upper() not in ("NV", "N/V"):
        query += f" {vintage}"

    # Correct JR tastings search URL — uses search-full parameter with quoted term
    encoded = urllib.parse.quote(f'"{query}"')
    return f"https://www.jancisrobinson.com/tastings?search-full={encoded}"


def main():
    if len(sys.argv) < 2:
        print("Usage: python verify_wines.py <path_to_xlsx>")
        print('Example: python verify_wines.py "C:/Users/You/Desktop/wines.xlsx"')
        sys.exit(1)

    path = sys.argv[1]
    print(f"Reading: {path}\n")

    wines = parse_xlsx(path)
    print(f"Found {len(wines)} wines\n")
    print(f"{'#':>4}  {'Vintage':>7}  {'LWIN11':>11}  {'Name / JR Search URL'}")
    print("─" * 120)

    for i, w in enumerate(wines, 1):
        url = jr_search_url(w["name"], w["vintage"])
        print(f"{i:>4}  {w['vintage'] or 'NV':>7}  {w['lwin'][:11] if w['lwin'] else '':>11}  {w['name']}")
        print(f"       {'':>7}  {'':>11}  {url}")
        print()

    print(f"\nTotal: {len(wines)} wines")
    print("\nPaste any URL above into your browser to verify the wine exists on JancisRobinson.com")


if __name__ == "__main__":
    main()

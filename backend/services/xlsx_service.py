"""
services/xlsx_service.py
========================
XLSX review enrichment service.

Flow:
  1. User uploads JR-format XLSX template
  2. parse_xlsx()  → extracts wine rows (name, vintage, lwin, row_idx)
  3. create_job()  → stores job in memory
  4. run_job()     → background thread: search JR for each wine
  5. fill_xlsx()   → generates filled XLSX bytes
  6. User downloads the result

Template format (from JR export template):
  Row 1:  Headers — Publisher, LWIN11, Product_Name, Vintage,
                     Critic_Name, Score, Drink_From, Drink_To, Review_Date, Review
  Row 2:  Instructions/examples (skipped)
  Row 3+: Wine data rows
"""
from __future__ import annotations

import io
import re
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from models.job_state_model import load_xlsx_job, list_xlsx_jobs, save_xlsx_job
from models.xlsx_file_model import (
    create_xlsx_file,
    delete_xlsx_file_record,
    get_xlsx_file,
    list_xlsx_files,
    update_xlsx_file,
)
from services.normalize_service import is_paywall_note, normalize_review, reset_note_tracking

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_STORAGE_DIR = BASE_DIR / "data" / "xlsx_files"
XLSX_ORIGINAL_DIR = XLSX_STORAGE_DIR / "original"
XLSX_OUTPUT_DIR = XLSX_STORAGE_DIR / "output"
XLSX_PREVIEW_ROWS = 50
INCOMPLETE_RETRY_ATTEMPTS = 3
INCOMPLETE_RETRY_SLEEP_SEC = 2.0

for _path in (XLSX_ORIGINAL_DIR, XLSX_OUTPUT_DIR):
    _path.mkdir(parents=True, exist_ok=True)


# ─── Column aliases ───────────────────────────────────────────────────────────
# Maps internal field names → list of substrings to look for in headers

_ALIASES: Dict[str, List[str]] = {
    "publisher":   ["publisher", "source"],
    "lwin":        ["lwin"],
    "name":        ["product_name", "wine_name", "name"],
    "vintage":     ["vintage"],
    "jr_search_url": ["jr_search_url", "url_on_the_jr", "jr_url"],
    "jr_wine_name": ["jr_wine_name", "wine_on_the_jr"],
    "jr_producer": ["jr_producer", "producer_on_jr"],
    "jr_appellation": ["jr_appellation", "appellation_on_jr"],
    "rp_search_url": ["rp_search_url", "url_on_the_rp", "rp_url"],
    "js_search_url": ["js_search_url", "jamessuckling_url", "url_on_the_js", "url_on_the_james_suckling", "js_url"],
    "js_tasting_note_id": ["js_tasting_note_id", "jamessuckling_id", "tasting_note_id_on_js", "js_id"],
    "dc_wine_name": ["dc_wine_name", "decanter_wine_name", "name_on_decanter"],
    "dc_search_url": ["dc_search_url", "decanter_search_url", "url_on_decanter_search_page"],
    "dc_review_url": ["dc_review_url", "decanter_review_url", "url_on_decanter_sperate_wine_page", "url_on_decanter_separate_wine_page"],
    "critic":      ["critic_name", "critic", "reviewer", "taster"],
    "score":       ["score"],
    "drink_from":  ["drink_fro", "drink_from", "from"],
    "drink_to":    ["drink_to"],
    "review_date": ["review_d", "review_date", "date_tasted", "date"],
    "review":      ["review", "tasting_note", "note"],
    "source_url":  ["source_url", "review_url", "url"],
}


def _col_idx(headers: List[str], field: str) -> Optional[int]:
    """Find column index matching a field by alias."""
    for alias in _ALIASES.get(field, [field]):
        for i, h in enumerate(headers):
            h_norm = h.lower().replace(" ", "_").replace("-", "_")
            if alias in h_norm:
                return i
    return None


# ─── Parse ────────────────────────────────────────────────────────────────────

def parse_xlsx(file_bytes: bytes) -> List[Dict]:
    """
    Parse JR-format XLSX template and return list of wine dicts.

    Each dict: { row_idx (1-based), name, vintage, lwin, lwin7, prefilled, ...optional source hints }
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Find header row (first row containing "product_name" or "lwin")
    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        row_low = [str(c or "").strip().lower().replace(" ", "_") for c in row]
        if any("product_name" in c or "lwin" in c for c in row_low):
            hdr_idx = i
            break

    raw_headers = [str(c or "").strip() for c in rows[hdr_idx]]
    lwin_col    = _col_idx(raw_headers, "lwin")
    name_col    = _col_idx(raw_headers, "name")
    vintage_col = _col_idx(raw_headers, "vintage")
    jr_search_url_col = _col_idx(raw_headers, "jr_search_url")
    jr_wine_name_col = _col_idx(raw_headers, "jr_wine_name")
    jr_producer_col = _col_idx(raw_headers, "jr_producer")
    jr_appellation_col = _col_idx(raw_headers, "jr_appellation")
    rp_search_url_col = _col_idx(raw_headers, "rp_search_url")
    js_search_url_col = _col_idx(raw_headers, "js_search_url")
    js_tasting_note_id_col = _col_idx(raw_headers, "js_tasting_note_id")
    dc_wine_name_col = _col_idx(raw_headers, "dc_wine_name")
    dc_search_url_col = _col_idx(raw_headers, "dc_search_url")
    dc_review_url_col = _col_idx(raw_headers, "dc_review_url")
    critic_col = _col_idx(raw_headers, "critic")
    score_col = _col_idx(raw_headers, "score")
    drink_from_col = _col_idx(raw_headers, "drink_from")
    drink_to_col = _col_idx(raw_headers, "drink_to")
    review_date_col = _col_idx(raw_headers, "review_date")
    review_col = _col_idx(raw_headers, "review")
    source_url_col = _col_idx(raw_headers, "source_url")

    def cell(row, col):
        if col is None or col >= len(row) or row[col] is None:
            return ""
        return str(row[col]).strip()

    def is_instruction_row(row_vals: List[str]) -> bool:
        joined = " | ".join(v.lower() for v in row_vals if v).strip()
        if not joined:
            return False
        markers = ("optional if", "e.g.", "yyyy", "text")
        return any(m in joined for m in markers)

    wines = []
    # Data starts after header; instruction rows are skipped dynamically.
    for offset, row in enumerate(rows[hdr_idx + 1:]):
        row_idx_1b = hdr_idx + 1 + offset + 1   # 1-based for openpyxl

        name_raw = cell(row, name_col)
        lwin_raw    = cell(row, lwin_col)
        vintage_raw = cell(row, vintage_col)
        jr_search_url = cell(row, jr_search_url_col)
        jr_wine_name = cell(row, jr_wine_name_col)
        jr_producer = cell(row, jr_producer_col)
        jr_appellation = cell(row, jr_appellation_col)
        rp_search_url = cell(row, rp_search_url_col)
        js_search_url = cell(row, js_search_url_col)
        js_tasting_note_id = cell(row, js_tasting_note_id_col)
        dc_wine_name = cell(row, dc_wine_name_col)
        dc_search_url = cell(row, dc_search_url_col)
        dc_review_url = cell(row, dc_review_url_col)
        row_vals = [cell(row, i) for i in range(len(raw_headers))]

        if is_instruction_row(row_vals):
            continue
        if not name_raw and not lwin_raw:
            continue
        if not name_raw or "optional" in name_raw.lower():
            continue

        # Clean name: strip trailing year and pack-size notes like "(Magnum)"
        name_clean = re.sub(r"\s*\(.*?\)\s*$", "", name_raw).strip()
        name_clean = re.sub(r",?\s*\b(19|20)\d{2}\b\s*$", "", name_clean).strip()

        # Determine vintage: column value → LWIN chars 7-11
        vintage = ""
        if vintage_raw and vintage_raw.upper() not in ("NV", "N/V", ""):
            m = re.search(r"\b(19|20)\d{2}\b", vintage_raw)
            if m:
                vintage = m.group(0)
        if not vintage and len(lwin_raw) >= 11:
            lv = lwin_raw[7:11]
            if lv.isdigit() and int(lv) > 1900:
                vintage = lv

        lwin7 = lwin_raw[:7]  if len(lwin_raw) >= 7  else lwin_raw
        lwin  = lwin_raw[:11] if len(lwin_raw) >= 11 else lwin_raw

        wine = {
            "row_idx": row_idx_1b,
            "name":    name_clean,
            "raw_name": name_raw,
            "vintage": vintage,
            "lwin":    lwin,
            "lwin7":   lwin7,
            "prefilled": any((
                cell(row, critic_col),
                cell(row, score_col),
                cell(row, drink_from_col),
                cell(row, drink_to_col),
                cell(row, review_date_col),
                cell(row, review_col),
                cell(row, source_url_col),
            )),
        }
        if (
            jr_search_url or jr_wine_name or jr_producer or jr_appellation
            or rp_search_url or js_search_url or js_tasting_note_id
            or dc_wine_name or dc_search_url or dc_review_url
        ):
            wine["search_hints"] = {
                "jr_search_url": jr_search_url,
                "jr_wine_name": jr_wine_name,
                "jr_producer": jr_producer,
                "jr_appellation": jr_appellation,
                "rp_search_url": rp_search_url,
                "jamessuckling_url": js_search_url,
                "js_tasting_note_id": js_tasting_note_id,
                "decanter_wine_name": dc_wine_name,
                "decanter_search_url": dc_search_url,
                "decanter_review_url": dc_review_url,
            }
        wines.append(wine)

    return wines


def _normalize_lwin_token(value: str | int | None) -> str:
    digits = re.sub(r"\D+", "", str(value or "").strip())
    if len(digits) >= 11:
        return digits[:11]
    if len(digits) >= 7:
        return digits[:7]
    return ""


def _parse_lwin_filter(raw_value: str | None) -> tuple[set[str], list[str]]:
    raw = str(raw_value or "").strip()
    if not raw:
        return set(), []

    values: set[str] = set()
    invalid: list[str] = []
    for part in re.split(r"[\s,;]+", raw):
        token = str(part or "").strip()
        if not token:
            continue
        normalized = _normalize_lwin_token(token)
        if not normalized:
            invalid.append(token)
            continue
        values.add(normalized)
    return values, invalid


def apply_lwin_filter(wines: List[Dict], raw_value: str | None) -> tuple[List[Dict], Dict]:
    raw = str(raw_value or "").strip()
    if not raw:
        return wines, {
            "enabled": False,
            "raw": "",
            "values": [],
            "invalid_values": [],
            "matched_values": [],
            "unmatched_values": [],
            "selected_rows": len(wines),
            "original_rows": len(wines),
        }

    values, invalid_values = _parse_lwin_filter(raw)
    if not values:
        raise ValueError("LWIN filter did not contain any valid LWIN7 or LWIN11 values.")

    matched_values: set[str] = set()
    filtered: list[Dict] = []
    for wine in wines:
        lwin11 = _normalize_lwin_token(wine.get("lwin"))
        lwin7 = _normalize_lwin_token(wine.get("lwin7"))
        if lwin11 in values or lwin7 in values:
            filtered.append(wine)
            if lwin11 in values:
                matched_values.add(lwin11)
            if lwin7 in values:
                matched_values.add(lwin7)

    if not filtered:
        raise ValueError("None of the requested LWIN values were found in this file.")

    unmatched_values = sorted(values - matched_values)
    return filtered, {
        "enabled": True,
        "raw": raw,
        "values": sorted(values),
        "invalid_values": invalid_values,
        "matched_values": sorted(matched_values),
        "unmatched_values": unmatched_values,
        "selected_rows": len(filtered),
        "original_rows": len(wines),
    }


def detect_source_from_template(file_bytes: bytes) -> Optional[str]:
    """
    Infer source from Publisher column values in the uploaded template.
    Returns source key or None when unknown.
    """
    if not OPENPYXL_OK:
        return None

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None

    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        row_low = [str(c or "").strip().lower().replace(" ", "_") for c in row]
        if any("publisher" in c or "product_name" in c or "lwin" in c for c in row_low):
            hdr_idx = i
            break

    raw_headers = [str(c or "").strip() for c in rows[hdr_idx]]
    pub_col = _col_idx(raw_headers, "publisher")
    if pub_col is None:
        return None

    for row in rows[hdr_idx + 1: hdr_idx + 8]:
        if pub_col >= len(row):
            continue
        pub = str(row[pub_col] or "").strip().lower()
        if not pub:
            continue
        if "robert parker" in pub or "wine advocate" in pub:
            return "robertparker"
        if "jancis" in pub:
            return "jancisrobinson"
        if "james suckling" in pub:
            return "jamessuckling"

    return None


def _safe_storage_name(name: str) -> str:
    raw = str(name or "upload.xlsx").strip() or "upload.xlsx"
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("._")
    return safe or "upload.xlsx"


def _read_file_bytes(path_str: str) -> bytes:
    return Path(path_str).read_bytes()


def _preview_wines(wines: List[Dict], limit: int = XLSX_PREVIEW_ROWS) -> List[Dict]:
    return [
        {
            "row_idx": w.get("row_idx"),
            "name": w.get("name"),
            "raw_name": w.get("raw_name"),
            "vintage": w.get("vintage"),
            "lwin": w.get("lwin"),
            "lwin7": w.get("lwin7"),
            "prefilled": bool(w.get("prefilled")),
        }
        for w in wines[:limit]
    ]


def _summarize_file_record(record: dict) -> dict:
    return {
        "file_id": record["file_id"],
        "original_name": record["original_name"],
        "stored_name": record["stored_name"],
        "source": record.get("source"),
        "status": record.get("status") or "pending",
        "size_bytes": int(record.get("size_bytes") or 0),
        "total_rows": int(record.get("total_rows") or 0),
        "prefilled_rows": int(record.get("prefilled_rows") or 0),
        "done_rows": int(record.get("done_rows") or 0),
        "found_rows": int(record.get("found_rows") or 0),
        "active_job_id": record.get("active_job_id"),
        "last_job_id": record.get("last_job_id"),
        "last_error": record.get("last_error"),
        "created_at": record.get("created_at"),
        "updated_at": record.get("updated_at"),
        "has_output": bool(record.get("output_path")),
    }


def build_export_policy(source_key: str, lwin_filter: Dict | None = None) -> Dict[str, bool]:
    """
    Shared XLSX export policy for every source.

    The filled workbook should keep the same row layout as the uploaded file so
    users can see found and not-found outcomes in-place. Keeping this in one
    helper prevents JR/JS flows from drifting apart again.
    """
    _ = source_key
    _ = lwin_filter or {}
    return {
        "preserve_all_rows": True,
    }


# ─── Fill ─────────────────────────────────────────────────────────────────────

def fill_xlsx(
    template_bytes: bytes,
    results: List[Dict],
    wines: Optional[List[Dict]] = None,
    preserve_all_rows: bool = False,
    qa_stats: Optional[Dict[str, int]] = None,
) -> bytes:
    """
    Fill review data into the template XLSX and return the bytes.

    Each result dict: { row_idx, found, critic_name, score_20,
                        drink_from, drink_to, date_tasted, note }
    Found rows highlighted green, not-found rows highlighted dark red.
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl not installed")

    wb = load_workbook(filename=io.BytesIO(template_bytes))
    ws = wb.active

    # Find header row to locate writable columns
    hdr_row_num = 1
    for row in ws.iter_rows(min_row=1, max_row=5):
        vals = [str(c.value or "").strip().lower() for c in row]
        if any("product" in v or "lwin" in v for v in vals):
            hdr_row_num = row[0].row
            break

    col_letter: Dict[str, str] = {}
    claimed: set = set()
    for cell in ws[hdr_row_num]:
        v = str(cell.value or "").strip()
        if not v or cell.column_letter in claimed:
            continue
        v_norm = v.lower().replace(" ", "_").replace("-", "_")
        for field, aliases in _ALIASES.items():
            if field not in col_letter and any(a in v_norm for a in aliases):
                col_letter[field] = cell.column_letter
                claimed.add(cell.column_letter)
                break  # one field per column — prevents "review" matching "review_date" header

    found_fill    = PatternFill("solid", fgColor="1a5c2e")   # visible green
    notfound_fill = PatternFill("solid", fgColor="7a1e1e")   # visible red

    WRITE_FIELDS = ["critic", "score", "drink_from", "drink_to", "review_date", "review", "source_url"]
    cleared_existing_bad_rows = _sanitize_existing_sheet_rows(ws, hdr_row_num, col_letter)
    if qa_stats is not None:
        qa_stats["existing_bad_rows_cleared"] = int(cleared_existing_bad_rows)

    for res in results:
        r = res["row_idx"]
        if res.get("found"):
            if col_letter.get("critic"):      ws[f"{col_letter['critic']}{r}"]      = res.get("critic_name") or ""
            if col_letter.get("score"):       ws[f"{col_letter['score']}{r}"]       = res.get("score_20")
            if col_letter.get("drink_from"):  ws[f"{col_letter['drink_from']}{r}"]  = res.get("drink_from") or ""
            if col_letter.get("drink_to"):    ws[f"{col_letter['drink_to']}{r}"]    = res.get("drink_to") or ""
            if col_letter.get("review_date"): ws[f"{col_letter['review_date']}{r}"] = res.get("date_tasted") or ""
            if col_letter.get("review"):      ws[f"{col_letter['review']}{r}"]      = res.get("note") or ""
            if col_letter.get("source_url"):  ws[f"{col_letter['source_url']}{r}"]  = res.get("source_url") or ""
            fill = found_fill
        else:
            if res.get("clear_existing"):
                if col_letter.get("critic"):      ws[f"{col_letter['critic']}{r}"]      = ""
                if col_letter.get("score"):       ws[f"{col_letter['score']}{r}"]       = ""
                if col_letter.get("drink_from"):  ws[f"{col_letter['drink_from']}{r}"]  = ""
                if col_letter.get("drink_to"):    ws[f"{col_letter['drink_to']}{r}"]    = ""
                if col_letter.get("review_date"): ws[f"{col_letter['review_date']}{r}"] = ""
                if col_letter.get("review"):      ws[f"{col_letter['review']}{r}"]      = ""
                if col_letter.get("source_url"):  ws[f"{col_letter['source_url']}{r}"]  = ""
            fill = notfound_fill

        for field in WRITE_FIELDS:
            cl = col_letter.get(field)
            if cl:
                ws[f"{cl}{r}"].fill = fill

    if not preserve_all_rows:
        row_map = {int(res["row_idx"]): res for res in results if res.get("row_idx")}
        keep_rows = set(range(1, hdr_row_num + 1))
        if wines:
            for wine in wines:
                row_idx = int(wine.get("row_idx") or 0)
                if not row_idx:
                    continue
                result = row_map.get(row_idx)
                if wine.get("prefilled") or (result and result.get("found")):
                    keep_rows.add(row_idx)
        _prune_empty_result_rows(ws, keep_rows)
    _remove_empty_placeholder_columns(ws, hdr_row_num)
    duplicate_lwin_rows_removed = _dedupe_rows_by_lwin(ws, hdr_row_num, col_letter)
    if qa_stats is not None:
        qa_stats["duplicate_lwin_rows_removed"] = int(duplicate_lwin_rows_removed)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def create_file_upload(
    filename: str,
    template_bytes: bytes,
    source_key: str,
) -> dict:
    wines = parse_xlsx(template_bytes)
    if not wines:
        raise ValueError("No wine rows found. Check the file format.")
    file_id = str(uuid.uuid4())
    stored_name = f"{file_id}_{_safe_storage_name(filename)}"
    original_path = XLSX_ORIGINAL_DIR / stored_name
    original_path.write_bytes(template_bytes)
    record = {
        "file_id": file_id,
        "original_name": filename or "upload.xlsx",
        "stored_name": stored_name,
        "original_path": str(original_path),
        "output_path": None,
        "source": source_key,
        "status": "pending",
        "size_bytes": len(template_bytes),
        "total_rows": len(wines),
        "prefilled_rows": sum(1 for w in wines if w.get("prefilled")),
        "done_rows": 0,
        "found_rows": 0,
        "active_job_id": None,
        "last_job_id": None,
        "last_error": None,
    }
    create_xlsx_file(record)
    return {
        "file_id": file_id,
        "wines": wines,
        "template_bytes": template_bytes,
        "record": record,
    }


def _prune_empty_result_rows(ws, keep_rows: set[int]) -> None:
    max_row = ws.max_row or 0
    for row_idx in range(max_row, 0, -1):
        if row_idx not in keep_rows:
            ws.delete_rows(row_idx, 1)


def _remove_empty_placeholder_columns(ws, hdr_row_num: int) -> None:
    removable = []
    for cell in ws[hdr_row_num]:
        header = str(cell.value or "").strip().lower()
        if header not in ("place", "place.1", "place.2"):
            continue
        has_data = False
        for row_idx in range(hdr_row_num + 1, (ws.max_row or hdr_row_num) + 1):
            if str(ws.cell(row=row_idx, column=cell.col_idx).value or "").strip():
                has_data = True
                break
        if not has_data:
            removable.append(cell.col_idx)
    for col_idx in reversed(removable):
        ws.delete_cols(col_idx, 1)


def _clear_row_scraped_cells(ws, row_idx: int, col_letter: Dict[str, str]) -> None:
    for field in ("critic", "score", "drink_from", "drink_to", "review_date", "review", "source_url"):
        cl = col_letter.get(field)
        if cl:
            ws[f"{cl}{row_idx}"] = ""


def _sanitize_existing_sheet_rows(ws, hdr_row_num: int, col_letter: Dict[str, str]) -> int:
    cleared = 0
    for row_idx in range(hdr_row_num + 1, (ws.max_row or hdr_row_num) + 1):
        note = str(ws[f"{col_letter['review']}{row_idx}"].value or "").strip() if col_letter.get("review") else ""
        score_value = ws[f"{col_letter['score']}{row_idx}"].value if col_letter.get("score") else None
        score_text = str(score_value).strip() if score_value is not None else ""
        drink_from = str(ws[f"{col_letter['drink_from']}{row_idx}"].value or "").strip() if col_letter.get("drink_from") else ""
        vintage_text = str(ws[f"{col_letter['vintage']}{row_idx}"].value or "").strip() if col_letter.get("vintage") else ""

        clear_row = False
        if note and is_paywall_note(note):
            clear_row = True
        elif note and not score_text:
            clear_row = True
        elif score_text and len(note) < 25:
            clear_row = True
        elif drink_from and vintage_text:
            try:
                drink_year = int(re.search(r"\b(19\d{2}|20\d{2})\b", drink_from).group(1))
                vintage_year = int(re.search(r"\b(19\d{2}|20\d{2})\b", vintage_text).group(1))
                if drink_year < vintage_year:
                    clear_row = True
            except Exception:
                pass

        if clear_row:
            _clear_row_scraped_cells(ws, row_idx, col_letter)
            cleared += 1
    return cleared


def _dedupe_rows_by_lwin(ws, hdr_row_num: int, col_letter: Dict[str, str]) -> int:
    lwin_col = col_letter.get("lwin")
    if not lwin_col:
        return 0
    duplicate_rows: list[int] = []
    best_by_lwin: Dict[str, tuple[int, tuple]] = {}

    for row_idx in range(hdr_row_num + 1, (ws.max_row or hdr_row_num) + 1):
        lwin_value = _normalize_lwin_token(ws[f"{lwin_col}{row_idx}"].value)
        if not lwin_value:
            continue
        note = ""
        if col_letter.get("review"):
            note = str(ws[f"{col_letter['review']}{row_idx}"].value or "").strip()
        score_value = None
        if col_letter.get("score"):
            score_value = ws[f"{col_letter['score']}{row_idx}"].value
        try:
            score_num = float(score_value) if score_value not in (None, "") else -9999.0
        except Exception:
            score_num = -9999.0
        date_value = ""
        if col_letter.get("review_date"):
            date_value = str(ws[f"{col_letter['review_date']}{row_idx}"].value or "").strip()
        row_key = (
            1 if len(note) >= 25 else 0,
            1 if score_num > -9999.0 else 0,
            score_num,
            date_value,
            -row_idx,
        )
        current = best_by_lwin.get(lwin_value)
        if current is None or row_key > current[1]:
            if current is not None:
                duplicate_rows.append(current[0])
            best_by_lwin[lwin_value] = (row_idx, row_key)
        else:
            duplicate_rows.append(row_idx)

    for row_idx in sorted(set(duplicate_rows), reverse=True):
        ws.delete_rows(row_idx, 1)
    return len(set(duplicate_rows))


def list_file_jobs(file_id: str) -> List[Dict]:
    return list_xlsx_jobs(file_id=file_id)


def list_file_records() -> List[Dict]:
    return [_summarize_file_record(row) for row in list_xlsx_files()]


def get_file_detail(file_id: str, include_preview: bool = True) -> Optional[Dict]:
    record = get_xlsx_file(file_id)
    if not record:
        return None

    detail = _summarize_file_record(record)
    if include_preview:
        try:
            template_bytes = _read_file_bytes(record["original_path"])
            wines = parse_xlsx(template_bytes)
            detail["preview_rows"] = _preview_wines(wines)
            detail["preview_count"] = min(len(wines), XLSX_PREVIEW_ROWS)
            detail["preview_deferred"] = False
        except Exception as e:
            detail["preview_rows"] = []
            detail["preview_count"] = 0
            detail["preview_error"] = str(e)
            detail["preview_deferred"] = False
    else:
        detail["preview_rows"] = []
        detail["preview_count"] = 0
        detail["preview_deferred"] = True

    jobs = list_file_jobs(file_id)
    detail["jobs"] = jobs
    active_job_id = record.get("active_job_id")
    last_job_id = record.get("last_job_id")
    detail["active_job"] = get_job(active_job_id) if active_job_id else None
    detail["last_job"] = get_job(last_job_id) if last_job_id else None
    if detail["active_job"] and detail["active_job"].get("status") in ("done", "stopped", "error"):
        detail["last_job"] = detail["last_job"] or detail["active_job"]
        detail["active_job"] = None
        update_xlsx_file(file_id, {"active_job_id": None})
    return detail


def get_file_download(file_id: str, kind: str = "original") -> Optional[Dict]:
    record = get_xlsx_file(file_id)
    if not record:
        return None
    if kind == "output":
        output_path = str(record.get("output_path") or "").strip()
        if not output_path:
            return None
        path = Path(output_path)
        if not path.exists():
            return None
        base = Path(record["original_name"]).stem
        return {
            "path": str(path),
            "download_name": f"maaike_{base}_filled.xlsx",
        }

    path = Path(record["original_path"])
    if not path.exists():
        return None
    return {
        "path": str(path),
        "download_name": record["original_name"],
    }


def delete_file(file_id: str) -> tuple[bool, str]:
    record = get_xlsx_file(file_id)
    if not record:
        return False, "not_found"

    active_job_id = record.get("active_job_id")
    active = load_xlsx_job(active_job_id) if active_job_id else None
    if active and active.get("status") in ("running", "pending"):
        return False, "job_running"

    for raw_path in (record.get("original_path"), record.get("output_path")):
        path_str = str(raw_path or "").strip()
        if not path_str:
            continue
        try:
            Path(path_str).unlink(missing_ok=True)
        except Exception:
            pass

    with _lock:
        if active_job_id and active_job_id in _jobs:
            _jobs.pop(active_job_id, None)
        last_job_id = record.get("last_job_id")
        if last_job_id and last_job_id in _jobs:
            _jobs.pop(last_job_id, None)

    delete_xlsx_file_record(file_id)
    return True, "ok"


# ─── Job management ───────────────────────────────────────────────────────────

_jobs: Dict[str, dict] = {}
_lock = threading.Lock()
_recovery_done = False


def _sync_file_record_from_job(job_id: str, job: dict) -> None:
    file_id = str(job.get("file_id") or "").strip()
    if not file_id:
        return
    patch = {
        "source": job.get("source"),
        "status": job.get("status") or "pending",
        "done_rows": int(job.get("done") or 0),
        "found_rows": int(job.get("found") or 0),
        "last_job_id": job_id,
        "last_error": job.get("error"),
    }
    if "output_path" in job:
        patch["output_path"] = job.get("output_path")
    if job.get("status") in ("running", "pending"):
        patch["active_job_id"] = job_id
    elif job.get("status") in ("done", "stopped", "error"):
        patch["active_job_id"] = None
    update_xlsx_file(file_id, patch)


def _snapshot_job(job_id: str, job: dict) -> None:
    try:
        save_xlsx_job(job_id, job)
        _sync_file_record_from_job(job_id, job)
    except Exception:
        pass


def _ensure_job_loaded(job_id: str) -> None:
    _recover_jobs_once()
    with _lock:
        if job_id in _jobs:
            return
    loaded = load_xlsx_job(job_id)
    if not loaded:
        return
    with _lock:
        _jobs[job_id] = loaded


def _recover_jobs_once() -> None:
    global _recovery_done
    if _recovery_done:
        return
    _recovery_done = True
    try:
        rows = list_xlsx_jobs(["running", "pending"])
        for row in rows:
            job_id = row.get("job_id")
            if not job_id:
                continue
            loaded = load_xlsx_job(job_id)
            if not loaded:
                continue
            loaded["status"] = "pending"
            loaded["stop_requested"] = False
            loaded["auto_stopped"] = False
            loaded["error"] = None
            with _lock:
                _jobs[job_id] = loaded
            _snapshot_job(job_id, loaded)
            threading.Thread(
                target=run_job,
                args=(job_id, loaded.get("source") or "jancisrobinson", float(loaded.get("sleep_sec") or 2.5)),
                daemon=True,
            ).start()
    except Exception:
        pass


def _format_progress_label(search_name: str, vintage: str) -> str:
    text = str(search_name or "").strip()
    year = str(vintage or "").strip()
    if not year:
        return text
    if re.search(rf"\b{re.escape(year)}\b", text):
        return text
    return f"{text} {year}".strip()


def create_job(
    template_bytes: bytes,
    wines: List[Dict],
    source_key: str = "jancisrobinson",
    sleep_sec: float = 2.5,
    start_item: int = 1,
    file_id: str | None = None,
    lwin_filter: Dict | None = None,
    preserve_all_rows: bool | None = None,
) -> str:
    _recover_jobs_once()
    total = len(wines)
    start_index = max(0, min(total, int(start_item or 1) - 1))
    initial_found = sum(1 for w in wines[:start_index] if w.get("prefilled"))
    export_policy = build_export_policy(source_key, lwin_filter)
    job_id = str(uuid.uuid4())
    with _lock:
        _jobs[job_id] = {
            "file_id":        file_id,
            "status":         "pending",
            "total":          total,
            "done":           start_index,
            "found":          initial_found,
            "stop_requested": False,
            "auto_stopped":   False,
            "source":         source_key,
            "sleep_sec":      float(sleep_sec),
            "start_item":     start_index + 1,
            "start_index":    start_index,
            "initial_found":  initial_found,
            "lwin_filter":    lwin_filter or {},
            "preserve_all_rows": (
                export_policy["preserve_all_rows"]
                if preserve_all_rows is None else bool(preserve_all_rows)
            ),
            "wines":          wines,
            "template_bytes": template_bytes,
            "results":        [],
            "output_bytes":   None,
            "error":          None,
            "log":            [],
        }
        _snapshot_job(job_id, _jobs[job_id])
    return job_id


def get_job(job_id: str) -> Optional[Dict]:
    _ensure_job_loaded(job_id)
    with _lock:
        j = _jobs.get(job_id)
        if not j:
            return None
        return {
            "job_id": job_id,
            "file_id": j.get("file_id"),
            "status": j["status"],
            "total":  j["total"],
            "done":   j["done"],
            "found":  j["found"],
            "start_item": int(j.get("start_item") or 1),
            "stop_requested": bool(j.get("stop_requested")),
            "auto_stopped": bool(j.get("auto_stopped")),
            "pct":    round(j["done"] / j["total"] * 100, 1) if j["total"] else 0,
            "ready":  j["output_bytes"] is not None,
            "error":  j["error"],
            "source": j.get("source"),
            "sleep_sec": j.get("sleep_sec"),
            "lwin_filter": j.get("lwin_filter") or {},
            "preserve_all_rows": bool(j.get("preserve_all_rows")),
            "log":    j["log"][-60:],
        }


def get_job_output(job_id: str) -> Optional[bytes]:
    _ensure_job_loaded(job_id)
    with _lock:
        j = _jobs.get(job_id)
        return j["output_bytes"] if j else None


def _job_output_path(job_id: str, job: dict) -> Optional[Path]:
    file_id = str(job.get("file_id") or "").strip()
    if not file_id:
        return None
    record = get_xlsx_file(file_id)
    if not record:
        return None
    output_name = f"{file_id}_{Path(record['original_name']).stem}_filled.xlsx"
    return XLSX_OUTPUT_DIR / _safe_storage_name(output_name)


def get_job_progress_download(job_id: str) -> Optional[Dict]:
    """
    Build a downloadable XLSX from the rows processed so far.

    This is intentionally usable while the job is still running, so users can
    save partial progress before waiting for the background thread to finish.
    """
    _ensure_job_loaded(job_id)
    with _lock:
        j = _jobs.get(job_id)
        if not j:
            return None
        template_bytes = j.get("template_bytes")
        results = list(j.get("results") or [])
        wines = list(j.get("wines") or [])
        preserve_all_rows = bool(j.get("preserve_all_rows"))

    if not template_bytes:
        return None

    output_bytes = fill_xlsx(
        template_bytes,
        results,
        wines,
        preserve_all_rows=preserve_all_rows,
    )
    output_path = _job_output_path(job_id, {"file_id": j.get("file_id")})
    if output_path:
        output_path.write_bytes(output_bytes)

    with _lock:
        current = _jobs.get(job_id)
        if current:
            current["output_bytes"] = output_bytes
            if output_path:
                current["output_path"] = str(output_path)
            _snapshot_job(job_id, current)

    if output_path:
        record = get_xlsx_file(str(j.get("file_id") or ""))
        base = Path(record["original_name"]).stem if record else "reviews"
        return {
            "path": str(output_path),
            "download_name": f"maaike_{base}_progress.xlsx",
        }
    return {
        "bytes": output_bytes,
        "download_name": "maaike_reviews_progress.xlsx",
    }


def _stop_requested(job_id: str) -> bool:
    with _lock:
        j = _jobs.get(job_id)
        return not j or bool(j.get("stop_requested")) or j.get("status") == "stopped"


def _sleep_or_stop(job_id: str, seconds: float) -> bool:
    remaining = max(0.0, float(seconds or 0.0))
    while remaining > 0:
        if _stop_requested(job_id):
            return True
        chunk = min(0.25, remaining)
        time.sleep(chunk)
        remaining -= chunk
    return _stop_requested(job_id)


def request_stop(job_id: str) -> bool:
    _ensure_job_loaded(job_id)
    should_build_progress = False
    with _lock:
        j = _jobs.get(job_id)
        if not j:
            return False
        j["stop_requested"] = True
        j["status"] = "stopped"
        j["done"] = int(j.get("start_index") or 0) + len(j.get("results") or [])
        j["found"] = int(j.get("initial_found") or 0) + sum(1 for r in (j.get("results") or []) if r.get("found"))
        j["error"] = None
        j.setdefault("log", []).append("Stop requested. Saved current progress for download.")
        should_build_progress = bool(j.get("template_bytes"))
        _snapshot_job(job_id, j)

    if should_build_progress:
        try:
            get_job_progress_download(job_id)
        except Exception:
            pass
    return True


def resume_job(job_id: str) -> tuple[bool, str]:
    _ensure_job_loaded(job_id)
    with _lock:
        j = _jobs.get(job_id)
        if not j:
            return False, "not_found"
        if j["status"] in ("running", "pending"):
            return False, "already_running"
        if int(j.get("done") or 0) >= int(j.get("total") or 0):
            return False, "already_done"

        j["status"] = "pending"
        j["stop_requested"] = False
        j["auto_stopped"] = False
        j["error"] = None
        j["output_bytes"] = None
        j["output_path"] = None
        _snapshot_job(job_id, j)
        return True, "ok"


def restart_file_job(
    file_id: str,
    source_key: str,
    sleep_sec: float = 2.5,
    start_item: int = 1,
    lwin_filter_raw: str | None = None,
) -> tuple[Optional[dict], str]:
    record = get_xlsx_file(file_id)
    if not record:
        return None, "not_found"

    active_job_id = record.get("active_job_id")
    active = load_xlsx_job(active_job_id) if active_job_id else None
    if active and active.get("status") in ("running", "pending"):
        return None, "already_running"

    old_output = str(record.get("output_path") or "").strip()
    if old_output:
        try:
            Path(old_output).unlink(missing_ok=True)
        except Exception:
            pass

    template_bytes = _read_file_bytes(record["original_path"])
    all_wines = parse_xlsx(template_bytes)
    wines, lwin_filter = apply_lwin_filter(all_wines, lwin_filter_raw)
    job_id = create_job(
        template_bytes,
        wines,
        source_key=source_key or (record.get("source") or "jancisrobinson"),
        sleep_sec=float(sleep_sec or 2.5),
        start_item=int(start_item or 1),
        file_id=file_id,
        lwin_filter=lwin_filter,
    )
    update_xlsx_file(file_id, {
        "source": source_key or (record.get("source") or "jancisrobinson"),
        "status": "pending",
        "done_rows": max(0, int(start_item or 1) - 1),
        "found_rows": 0,
        "active_job_id": job_id,
        "last_job_id": job_id,
        "last_error": None,
        "output_path": None,
    })
    return {
        "job_id": job_id,
        "total": len(wines),
        "file_total": len(all_wines),
        "start_item": int(start_item or 1),
        "source": source_key or (record.get("source") or "jancisrobinson"),
        "file_id": file_id,
        "lwin_filter": lwin_filter,
    }, "ok"


# ─── Background runner ────────────────────────────────────────────────────────

def _parse_date(dt_str: str) -> Optional[datetime]:
    """Parse review date string into datetime for sorting."""
    for fmt in ("%d %b %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(dt_str.strip(), fmt)
        except (ValueError, AttributeError):
            pass
    return None


def _has_real_note(review: Dict) -> bool:
    note = str(review.get("note") or review.get("tasting_note") or "").strip()
    return len(note) >= 25


def _has_score(review: Dict) -> bool:
    return any(
        review.get(key) is not None
        for key in ("score_20", "score_100", "score_native", "score")
    )


def _quality_flags(review: Dict) -> set[str]:
    return {str(flag).strip() for flag in (review.get("_quality_flags") or []) if str(flag).strip()}


def _choose_best_review(reviews: List[Dict]) -> tuple[Optional[Dict], List[str]]:
    if not reviews:
        return None, ["no_reviews"]

    scored_complete = [r for r in reviews if _has_score(r) and _has_real_note(r)]
    if scored_complete:
        scored_complete.sort(key=_review_pick_sort_key, reverse=True)
        return scored_complete[0], []

    scored_only = [r for r in reviews if _has_score(r)]
    if scored_only:
        scored_only.sort(key=_review_pick_sort_key, reverse=True)
        best = scored_only[0]
        reasons = sorted(_quality_flags(best) | {"score_only"})
        return best, reasons

    note_only = [r for r in reviews if _has_real_note(r)]
    if note_only:
        note_only.sort(key=_review_pick_sort_key, reverse=True)
        best = note_only[0]
        reasons = sorted(_quality_flags(best) | {"note_without_score"})
        return best, reasons

    fallback = list(reviews)
    fallback.sort(key=_review_pick_sort_key, reverse=True)
    best = fallback[0]
    reasons = sorted(_quality_flags(best) | {"empty_review"})
    return best, reasons


def _review_pick_sort_key(review: Dict) -> tuple:
    has_note = 1 if _has_real_note(review) else 0
    has_score = 1 if _has_score(review) else 0
    parsed_date = _parse_date(str(review.get("date_tasted") or review.get("review_date") or "")) or datetime.min
    score = review.get("score_20")
    if score is None:
        score = review.get("score_100")
    if score is None:
        score = review.get("score_native")
    if score is None:
        score = review.get("score")
    try:
        score_num = float(score) if score is not None else -9999.0
    except Exception:
        score_num = -9999.0
    return (has_note, has_score, parsed_date, score_num)


def _is_network_error(exc: Exception) -> bool:
    """Best-effort detector for transient connectivity errors from source scrapers."""
    if isinstance(exc, (ConnectionError, TimeoutError, OSError)):
        return True
    name = type(exc).__name__.lower()
    text = str(exc).lower()
    markers = (
        "connectionerror",
        "newconnectionerror",
        "maxretryerror",
        "proxyerror",
        "name or service not known",
        "temporary failure in name resolution",
        "failed to establish a new connection",
        "read timed out",
        "connect timeout",
        "network is unreachable",
        "dns",
    )
    return any(m in name or m in text for m in markers)


def _is_access_blocked_error(exc: Exception) -> bool:
    text = str(exc).lower()
    markers = (
        "blocked by cloudflare",
        "cf_clearance",
        "cloudflare challenge",
        "access blocked",
        "unauthorized",
        "forbidden",
        "expired",
        "cookie",
        "cookies",
        "session token",
        "no session",
        "login",
        "log in",
        "sign in",
        "members only",
        "locked content",
        "teaser or locked content",
    )
    return any(marker in text for marker in markers)


def _result_note_group_key(result: Dict) -> str:
    note = str(result.get("note") or "").strip()
    if len(note) < 40:
        return ""
    reviewer = str(result.get("critic_name") or "").strip().lower()
    date_tasted = str(result.get("date_tasted") or "").strip()
    if not reviewer and not date_tasted:
        return ""
    fp = note[:120].lower().replace(" ", "").replace(",", "")
    return f"{fp}|{date_tasted}|{reviewer}"


def _strip_duplicate_result_notes(results: List[Dict]) -> int:
    groups: Dict[str, List[int]] = {}
    for idx, result in enumerate(results):
        if not result.get("found"):
            continue
        key = _result_note_group_key(result)
        if not key:
            continue
        groups.setdefault(key, []).append(idx)

    cleared = 0
    for indexes in groups.values():
        wine_keys = {
            (
                str(results[i].get("_lwin") or "").strip(),
                str(results[i].get("_wine_name") or "").strip().lower(),
                str(results[i].get("_vintage") or "").strip(),
            )
            for i in indexes
        }
        if len(indexes) < 2 or len(wine_keys) < 2:
            continue
        for i in indexes:
            note = str(results[i].get("note") or "").strip()
            if note:
                results[i]["note"] = ""
                cleared += 1
    return cleared


def _invalidate_incomplete_result(result: Dict, reasons: List[str] | set[str]) -> None:
    reason_list = [str(r).strip() for r in reasons if str(r).strip()]
    result["found"] = False
    result["clear_existing"] = True
    result["critic_name"] = ""
    result["score_20"] = ""
    result["drink_from"] = ""
    result["drink_to"] = ""
    result["date_tasted"] = ""
    result["note"] = ""
    result["source_url"] = ""
    result["_score_20_db"] = None
    result["_score_100_db"] = None
    result["_quality_reasons"] = sorted(set(reason_list))


def _finalize_result_quality(results: List[Dict]) -> int:
    invalidated = 0
    fatal_reasons = {
        "duplicate_note",
        "missing_note",
        "missing_score",
        "paywall_note",
        "invalid_drink_from",
        "invalid_drink_to",
        "reversed_drink_window",
    }
    for result in results:
        if not result.get("found"):
            continue
        note = str(result.get("note") or "").strip()
        has_score = any(
            result.get(key) is not None and str(result.get(key)).strip() != ""
            for key in ("_score_20_db", "_score_100_db", "score_20")
        )
        reasons = set(result.get("_quality_reasons") or [])
        if len(note) < 25:
            reasons.add("missing_note")
        if not has_score:
            reasons.add("missing_score")
        if reasons & fatal_reasons:
            _invalidate_incomplete_result(result, reasons)
            invalidated += 1
    return invalidated


def _save_found_results_to_db(results: List[Dict], source_key: str, job_id: str) -> int:
    from models.wine_model import upsert_review_wine

    saved = 0
    for result in results:
        if not result.get("found"):
            continue
        upsert_review_wine(
            result.get("_lwin") or "",
            {
                "name": result.get("_wine_name") or "",
                "vintage": result.get("_vintage") or None,
            },
            {
                "source": source_key,
                "score_20": result.get("_score_20_db"),
                "score_100": result.get("_score_100_db"),
                "reviewer": result.get("critic_name") or None,
                "note": result.get("note") or "",
                "date_tasted": result.get("date_tasted") or None,
                "drink_from": result.get("drink_from") or None,
                "drink_to": result.get("drink_to") or None,
                "review_url": result.get("source_url") or None,
            },
            upload_batch=f"xlsx_{job_id[:8]}",
        )
        saved += 1
    return saved


def _fetch_reviews_for_attempt(
    source_key: str,
    session,
    search_name: str,
    vintage: str,
    lwin: str,
    sleep_sec: float,
    search_hints: Dict | None,
):
    if source_key == "jancisrobinson":
        from maaike_phase1 import search_wine
        return search_wine(session, search_name, vintage, lwin, search_hints)
    from services.enrich_service import _search_source
    return _search_source(source_key, session, search_name, vintage, lwin, sleep_sec, search_hints)


def _select_candidate_review(
    source_key: str,
    reviews: List[Dict],
    wine_name: str,
    vintage: str,
    source_label: str,
    log,
) -> tuple[Optional[Dict], List[str]]:
    raw_paywall_notes = [
        str((review.get("note") or review.get("tasting_note") or "")).strip()
        for review in reviews
        if is_paywall_note(str((review.get("note") or review.get("tasting_note") or "")).strip())
    ]
    normalized_reviews = [
        normalize_review(
            source_key,
            review,
            wine_name=wine_name,
            wine_vintage=vintage,
            flag_duplicate_notes=True,
        )
        for review in reviews
    ]
    if raw_paywall_notes and not any(_has_real_note(r) for r in normalized_reviews):
        log(
            f"  - {source_label} returned a locked/teaser note; "
            "keeping score metadata and leaving note blank"
        )

    normalized_reviews.sort(key=_review_pick_sort_key, reverse=True)
    rev, quality_reasons = _choose_best_review(normalized_reviews)
    if not rev:
        return None, ["no_reviews"]

    score_20 = rev.get("score_20")
    score_100 = rev.get("score_100")
    if score_100 is None:
        score_100 = rev.get("score_native")
    if score_100 is None:
        score_100 = rev.get("score")
    out_score = score_20 if score_20 is not None else score_100

    reasons = set(quality_reasons or [])
    fatal_reasons = {
        "paywall_note",
        "invalid_drink_from",
        "invalid_drink_to",
        "reversed_drink_window",
    }
    if out_score is None:
        reasons.add("missing_score")
    if not _has_real_note(rev):
        reasons.add("missing_note")
    if reasons & fatal_reasons:
        reasons.add("bad_row_data")
    return rev, sorted(reasons)


def run_job(job_id: str, source_key: str = "jancisrobinson", sleep_sec: float = 2.5):
    """
    Background thread: search the selected source for each wine and fill results.
    """
    from services.session_service import clear_session, get_session
    _ensure_job_loaded(job_id)

    with _lock:
        j = _jobs.get(job_id)
        if not j:
            return
        if j.get("status") == "running":
            return
        if j.get("stop_requested") or j.get("status") == "stopped":
            return
        wines = list(j["wines"])
        template_bytes = j["template_bytes"]
        source_key = source_key or (j.get("source") or "jancisrobinson")
        sleep_sec = float(sleep_sec or j.get("sleep_sec") or 2.5)
        results: List[Dict] = list(j.get("results") or [])
        start_index = int(j.get("start_index") or 0)
        initial_found = int(j.get("initial_found") or 0)
        j["status"] = "running"
        j["source"] = source_key
        j["sleep_sec"] = float(sleep_sec)
        j["error"] = None
        j["stop_requested"] = False
        j["auto_stopped"] = False
        j["output_bytes"] = None
        j["done"] = start_index + len(results)
        j["found"] = initial_found + sum(1 for r in results if r.get("found"))
        _snapshot_job(job_id, j)

    source_label = source_key
    try:
        from config.sources import SOURCES
        source_label = SOURCES.get(source_key, {}).get("short") or source_key
    except Exception:
        pass

    def log(msg: str):
        with _lock:
            if job_id in _jobs:
                _jobs[job_id]["log"].append(msg)
                _snapshot_job(job_id, _jobs[job_id])

    session = get_session(source_key)
    if not session:
        with _lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = f"No session for '{source_key}' - upload cookies in Settings first"
            _snapshot_job(job_id, _jobs[job_id])
        return

    processed = start_index + len(results)
    saved = 0
    stopped = False
    auto_stopped = False
    consecutive_net_errors = 0
    qa_stats = {
        "retried_incomplete_rows": 0,
        "retry_attempts": 0,
        "rejected_incomplete_rows": 0,
        "invalidated_results": 0,
        "duplicate_notes_cleared": 0,
        "duplicate_lwin_rows_removed": 0,
        "existing_bad_rows_cleared": 0,
    }
    reset_note_tracking()

    if processed > 0:
        log(f"Resuming from row {processed + 1}/{len(wines)}...")

    for idx, wine in enumerate(wines[processed:], start=processed + 1):
        if _stop_requested(job_id):
            stopped = True
            log("Stop requested. Finishing current progress for download...")
            break

        name = wine["name"]
        search_name = wine.get("raw_name") if source_key == "decanter" and wine.get("raw_name") else name
        vintage = wine["vintage"]
        lwin = wine["lwin"]

        log(f"[{idx}/{len(wines)}] {_format_progress_label(search_name, vintage or 'NV')}")

        try:
            selected_review = None
            selected_reasons: List[str] = []
            selected_scores = (None, None, None)
            reviews = []
            for attempt in range(1, INCOMPLETE_RETRY_ATTEMPTS + 1):
                if _stop_requested(job_id):
                    stopped = True
                    break
                reviews = _fetch_reviews_for_attempt(
                    source_key,
                    session,
                    search_name,
                    vintage,
                    lwin,
                    sleep_sec,
                    wine.get("search_hints"),
                )
                if _stop_requested(job_id):
                    stopped = True
                    break
                if not reviews:
                    selected_review = None
                    selected_reasons = ["not_found"]
                    break

                rev, quality_reasons = _select_candidate_review(
                    source_key,
                    reviews,
                    name,
                    vintage,
                    source_label,
                    log,
                )
                if rev:
                    reviewer = rev.get("reviewer") or rev.get("critic_name") or ""
                    note = rev.get("note") or rev.get("tasting_note") or ""
                    date_tasted = rev.get("date_tasted") or rev.get("review_date") or ""
                    score_20 = rev.get("score_20")
                    score_100 = rev.get("score_100")
                    if score_100 is None:
                        score_100 = rev.get("score_native")
                    if score_100 is None:
                        score_100 = rev.get("score")
                    out_score = score_20 if score_20 is not None else score_100
                    selected_review = {
                        "row_idx": wine["row_idx"],
                        "found": True,
                        "critic_name": reviewer,
                        "score_20": out_score,
                        "drink_from": rev.get("drink_from") or "",
                        "drink_to": rev.get("drink_to") or "",
                        "date_tasted": date_tasted,
                        "note": note,
                        "source_url": rev.get("review_url") or rev.get("source_url") or "",
                        "_wine_name": name,
                        "_vintage": vintage,
                        "_lwin": lwin,
                        "_score_20_db": score_20,
                        "_score_100_db": score_100,
                        "_quality_reasons": list(quality_reasons or []),
                    }
                    selected_reasons = list(quality_reasons or [])
                    selected_scores = (out_score, score_20, score_100)
                    if out_score is not None and _has_real_note(rev):
                        break

                if attempt < INCOMPLETE_RETRY_ATTEMPTS:
                    qa_stats["retried_incomplete_rows"] += 1 if attempt == 1 else 0
                    qa_stats["retry_attempts"] += 1
                    log(
                        f"  - incomplete {source_label} result, retry {attempt}/{INCOMPLETE_RETRY_ATTEMPTS - 1} in {INCOMPLETE_RETRY_SLEEP_SEC:.0f}s"
                    )
                    if _sleep_or_stop(job_id, INCOMPLETE_RETRY_SLEEP_SEC):
                        stopped = True
                        break

            if stopped:
                log("Stop requested. Current in-flight row was not marked done.")
                break

            if selected_review and selected_scores[0] is not None and len(str(selected_review.get('note') or '').strip()) >= 25:
                consecutive_net_errors = 0
                results.append(selected_review)
                scale = "/20" if selected_scores[1] is not None else "/100"
                score_log = selected_scores[0] if selected_scores[0] is not None else "?"
                log(f"  + score {score_log}{scale} | {selected_review.get('date_tasted') or '-'}")
            else:
                consecutive_net_errors = 0
                if reviews:
                    qa_stats["rejected_incomplete_rows"] += 1
                    log(
                        f"  - rejected incomplete {source_label} result after {INCOMPLETE_RETRY_ATTEMPTS} attempts: "
                        f"{', '.join(selected_reasons or ['missing_data'])}"
                    )
                else:
                    log(f"  - not found on {source_label}")
                results.append({
                    "row_idx": wine["row_idx"],
                    "found": False,
                    "clear_existing": bool(reviews),
                    "_quality_reasons": list(selected_reasons or []),
                })

        except Exception as e:
            log(f"  x error: {type(e).__name__}: {e}")
            if _is_access_blocked_error(e):
                clear_session(source_key)
                auto_stopped = True
                stopped = True
                if source_key == "jancisrobinson":
                    log("Auto-stop: JR access is blocked by Cloudflare. Upload fresh browser cookies including cf_clearance, then Resume.")
                else:
                    log(f"Auto-stop: {source_label} session/cookies look invalid. Refresh cookies, then Resume.")
                log("Current row was not marked done, so Resume will retry this same row.")
            elif _is_network_error(e):
                consecutive_net_errors += 1
                log(f"  - network error streak: {consecutive_net_errors}")
                auto_stopped = True
                stopped = True
                log("Auto-stop: network unavailable. Resume when internet is back.")
                log("Current row was not marked done, so Resume will retry this same row.")
            else:
                consecutive_net_errors = 0
                results.append({"row_idx": wine["row_idx"], "found": False, "clear_existing": False})

        with _lock:
            if job_id in _jobs:
                _jobs[job_id]["done"] = start_index + len(results)
                _jobs[job_id]["found"] = initial_found + sum(1 for r in results if r.get("found"))
                _jobs[job_id]["results"] = list(results)
                if _jobs[job_id].get("stop_requested"):
                    stopped = True
                _snapshot_job(job_id, _jobs[job_id])
        if stopped:
            if auto_stopped:
                log("Generating partial XLSX after auto-stop...")
            else:
                log("Stop requested. Generating partial XLSX...")
            break

        if _sleep_or_stop(job_id, sleep_sec):
            stopped = True
            log("Stop requested. Generating partial XLSX...")
            break

    duplicate_notes_cleared = _strip_duplicate_result_notes(results)
    if duplicate_notes_cleared:
        qa_stats["duplicate_notes_cleared"] = duplicate_notes_cleared
        log(f"Cleared {duplicate_notes_cleared} duplicate session note(s) from XLSX results.")
        for result in results:
            if result.get("found") and len(str(result.get("note") or "").strip()) < 25:
                reasons = set(result.get("_quality_reasons") or [])
                reasons.add("duplicate_note")
                _invalidate_incomplete_result(result, reasons)
        log("Invalidated duplicate-note rows that became incomplete after cleanup.")

    invalidated_results = _finalize_result_quality(results)
    if invalidated_results:
        qa_stats["invalidated_results"] = invalidated_results
        log(f"Removed {invalidated_results} incomplete result(s) before DB save and XLSX export.")

    try:
        saved = _save_found_results_to_db(results, source_key, job_id)
    except Exception as db_err:
        log(f"  - DB save error: {type(db_err).__name__}: {db_err}")

    log("Generating filled XLSX...")
    try:
        output_bytes = fill_xlsx(
            template_bytes,
            results,
            wines,
            preserve_all_rows=bool(j.get("preserve_all_rows")),
            qa_stats=qa_stats,
        )
    except Exception as e:
        with _lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = f"Failed to generate XLSX: {e}"
            _snapshot_job(job_id, _jobs[job_id])
        return

    output_path = None
    file_id = str(j.get("file_id") or "").strip() if 'j' in locals() else ""
    if file_id:
        record = get_xlsx_file(file_id)
        if record:
            output_name = f"{file_id}_{Path(record['original_name']).stem}_filled.xlsx"
            output_path = XLSX_OUTPUT_DIR / _safe_storage_name(output_name)
            output_path.write_bytes(output_bytes)

    total = len(wines)
    found = initial_found + sum(1 for r in results if r.get("found"))
    pct = round(found / total * 100, 1) if total else 0
    log(f"Saved to DB: {saved}/{found} found review(s)")
    log(
        "QA Summary: "
        f"retried_rows={qa_stats['retried_incomplete_rows']} | "
        f"retry_attempts={qa_stats['retry_attempts']} | "
        f"rejected_incomplete={qa_stats['rejected_incomplete_rows']} | "
        f"invalidated_after_cleanup={qa_stats['invalidated_results']} | "
        f"duplicate_notes_cleared={qa_stats['duplicate_notes_cleared']} | "
        f"duplicate_lwin_rows_removed={qa_stats['duplicate_lwin_rows_removed']} | "
        f"existing_bad_rows_cleared={qa_stats['existing_bad_rows_cleared']}"
    )
    if stopped or _stop_requested(job_id):
        stopped = True
        if auto_stopped:
            log(f"Auto-stopped - processed {start_index + len(results)}/{total}, found {found}. Click Download or Resume.")
        else:
            log(f"Stopped - processed {start_index + len(results)}/{total}, found {found}. Click Download.")
    else:
        log(f"Done - {found}/{total} found ({pct}%). Click Download.")

    with _lock:
        _jobs[job_id]["status"] = "stopped" if stopped else "done"
        _jobs[job_id]["output_bytes"] = output_bytes
        _jobs[job_id]["done"] = start_index + len(results)
        _jobs[job_id]["found"] = found
        _jobs[job_id]["results"] = list(results)
        _jobs[job_id]["auto_stopped"] = auto_stopped
        if output_path:
            _jobs[job_id]["output_path"] = str(output_path)
        if auto_stopped and not _jobs[job_id].get("error"):
            if consecutive_net_errors > 0:
                _jobs[job_id]["error"] = "Network unavailable during row fetch. Resume will retry the same row."
            else:
                _jobs[job_id]["error"] = f"{source_label} session/cookies expired or blocked. Resume will retry the same row after cookies are refreshed."
        elif not auto_stopped:
            _jobs[job_id]["error"] = None
        _snapshot_job(job_id, _jobs[job_id])
    if file_id:
        update_xlsx_file(file_id, {
            "output_path": str(output_path) if output_path else None,
            "status": "stopped" if stopped else "done",
            "done_rows": start_index + len(results),
            "found_rows": found,
            "active_job_id": None,
            "last_job_id": job_id,
            "last_error": _jobs.get(job_id, {}).get("error"),
            "source": source_key,
        })

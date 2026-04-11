"""
test_matching.py
================
Run from the backend/ directory:

    python test_matching.py              # logic tests only (no network)
    python test_matching.py live         # + live RP search (needs cookies)
    python test_matching.py live jr      # + live JR search (needs cookies)

Tests:
  1. Name-match logic — RP  (_name_matches)
  2. Name-match logic — JR  (_name_matches_jr)
  3. Query builder  — RP    (_build_rp_queries)
  4. Live RP search  (optional, needs cookies/robertparker.json)
  5. Live JR search  (optional, needs cookies/jancisrobinson.json or real_cookies.json)
"""
import io
import os
import sys
import maaike_phase1 as jr_mod
from openpyxl import Workbook

# ── Import from both scrapers ──────────────────────────────────────────────────
from sources.robertparker import _name_matches, _name_tokens, _build_rp_queries
from sources.decanter import (
    _candidate_review_pool as _dc_candidate_review_pool,
    _build_search_queries as _build_dc_queries,
    _direct_detail_urls as _dc_direct_detail_urls,
    _manual_search_urls as _dc_manual_search_urls,
    _matches_query as _dc_matches_query,
    _openai_fallback_enabled as _dc_openai_fallback_enabled,
    _parse_detail_page as _dc_parse_detail_page,
    _preferred_query_name as _dc_preferred_query_name,
    _sort_candidates as _dc_sort_candidates,
)
from maaike_phase1 import (
    _build_search_queries,
    _candidate_prerank_score,
    _name_matches_jr,
    _name_tokens as _jr_tokens,
    _clean_name_variants,
    _parse_query_structured,
    _jr_query_identity,
    _jr_plausible_match,
    _jr_candidate_passes,
    _strict_match_structured,
)
from services import xlsx_service as xlsx_mod

PASS = "  ✓"
FAIL = "  ✗"

def check(label: str, result: bool, expected: bool):
    ok = result == expected
    mark = PASS if ok else FAIL
    detail = "" if ok else f"  ← expected {expected}, got {result}"
    print(f"{mark}  {label}{detail}")
    return ok


# ─── 1. RP name matching ──────────────────────────────────────────────────────
print("\n══ 1. RP name matching (_name_matches) ══════════════════════════════")

CASES_RP = [
    # (query,                                      hit_name,                                      expected)
    ("Charles Noellat, Nuits-Saint-Georges, Blanc", "Charles Noellat Nuits-Saint-Georges Blanc",    True),
    ("Charles Noellat, Nuits-Saint-Georges, Blanc", "Charles Noellat Richebourg",                   False),
    ("Charles Noellat, Nuits-Saint-Georges, Blanc", "Domaine de la Romanée-Conti Romanée-Conti",    False),
    ("Château Margaux, 2015",                        "Château Margaux",                             True),
    ("Château Margaux, 2015",                        "Château Palmer",                              False),
    ("Giacomo Conterno, Barolo, Monfortino Riserva", "Giacomo Conterno Barolo Monfortino Riserva",  True),
    ("Giacomo Conterno, Barolo, Monfortino Riserva", "Giacomo Conterno Barolo Francia",             False),
    ("Pétrus",                                       "Pétrus",                                      True),
    ("Pétrus",                                       "Le Pin",                                      False),
    ("Domaine Leroy, Chambolle-Musigny, Les Charmes","Leroy Chambolle-Musigny Les Charmes",         True),
    ("Domaine Leroy, Chambolle-Musigny, Les Charmes","Leroy Gevrey-Chambertin",                     False),
]

rp_ok = sum(check(f"query={q!r:50s} | hit={h!r:45s}", _name_matches(q, h), exp)
            for q, h, exp in CASES_RP)
print(f"\n  {rp_ok}/{len(CASES_RP)} passed")


# ─── 2. JR name matching ──────────────────────────────────────────────────────
print("\n══ 2. JR name matching (_name_matches_jr) ═══════════════════════════")

CASES_JR = [
    ("Charles Noellat, Nuits-Saint-Georges, Blanc 2006", "Charles Noellat, Nuits-Saint-Georges Blanc", True),
    ("Charles Noellat, Nuits-Saint-Georges, Blanc 2006", "Charles Noellat, Richebourg",               False),
    ("Château Latour 2010",                              "Château Latour",                            True),
    ("Château Latour 2010",                              "Château Léoville-Las Cases",                False),
    ("Domaine Leflaive, Puligny-Montrachet, Pucelles",   "Leflaive Puligny-Montrachet Pucelles",      True),
    ("Domaine Leflaive, Puligny-Montrachet, Pucelles",   "Leflaive Bâtard-Montrachet",                False),
]

CASES_JR += [
    ("Chateau Tour de Pez, Les Hauts de Pez, Saint-Estephe", "L'Ame du Domaine",                      False),
    ("Chateau Tour de Pez, Les Hauts de Pez, Saint-Estephe", "Tour de Pez Les Hauts de Pez",          True),
    ("Azienda Agricola Salvioni La Cerbaiola, Brunello di Montalcino", "Dry Amber Reserve",           False),
    ("Azienda Agricola Salvioni La Cerbaiola, Brunello di Montalcino", "Salvioni Brunello",           True),
    ("Azienda Agricola Salvioni La Cerbaiola, Brunello di Montalcino", "Salvioni La Cerbaiola",       True),
    ("Domaine d'Eugenie, Vosne-Romanee Premier Cru, Aux Brulees", "Aux Brulees",                      True),
    ("Chateau La Fleur de Bouard, Lalande de Pomerol", "La Fleur de Bouard",                          True),
    ("Chateau La Fleur de Bouard, Lalande de Pomerol", "Random Lalande de Pomerol",                   False),
]

jr_ok = sum(check(f"query={q!r:55s} | hit={h!r:45s}", _name_matches_jr(q, h), exp)
            for q, h, exp in CASES_JR)
print(f"\n  {jr_ok}/{len(CASES_JR)} passed")

print("\n══ 2b. JR variant generation (_clean_name_variants) ═════════════════════")
_v = _clean_name_variants("Chateau La Fleur de Bouard, Lalande de Pomerol")
has_bad = any("Ateau " in x for x in _v)
print(f"  {'✓' if not has_bad else '✗'}  no broken 'Ateau ...' variant")
check("JR emits Ch abbreviation", "Ch La Fleur De Bouard, Lalande De Pomerol" in _v, True)

_v2 = _clean_name_variants("Domaine d'Eugenie, Vosne-Romanee Premier Cru, Aux Brulees")
check("JR keeps apostrophe-aware producer form", "D'Eugenie Aux Brulees" in _v2, True)
check("JR emits decontracted producer form", "Eugenie Aux Brulees" in _v2, True)

_v3 = _clean_name_variants("Azienda Agricola Salvioni La Cerbaiola, Brunello di Montalcino")
check("JR emits bare left-side identity for 2-part names", "Salvioni La Cerbaiola" in _v3, True)

_v4 = _clean_name_variants("Georges Lignier Et Fils, Gevrey-Chambertin")
check("JR trims producer suffixes before region", "Georges Lignier" in _v4, True)

print("\n══ 2c. JR query identity (_jr_query_identity) ═══════════════════════════")
identity = _jr_query_identity("Domaine Leflaive, Puligny-Montrachet, Pucelles 2020", "2020")
check("producer parsed", identity.get("producer") == "Domaine Leflaive", True)
check("specific parsed", identity.get("specific") == "Pucelles", True)
check("vintage parsed", identity.get("vintage") == "2020", True)

print("\n══ 2d. JR structured query + variants ════════════════════════════════════")
structured = _parse_query_structured("Domaine Leflaive, Puligny-Montrachet, Pucelles 2020", "2020")
check("structured producer", structured.get("producer") == "Domaine Leflaive", True)
check("structured wine", structured.get("wine") == "Puligny-Montrachet", True)
check("structured appellation", structured.get("appellation") == "Pucelles", True)
structured_two_part = _parse_query_structured("Chateau La Fleur de Bouard, Lalande de Pomerol", "2020")
check("2-part region parsed as producer", structured_two_part.get("producer") == "Chateau La Fleur de Bouard", True)
check("2-part region parsed as appellation", structured_two_part.get("appellation") == "Lalande de Pomerol", True)
queries = _build_search_queries("Domaine Leflaive, Puligny-Montrachet, Pucelles", "2020", "LWIN12345672020", {
    "jr_search_url": "https://www.jancisrobinson.com/tastings?search-full=%22Leflaive%20Pucelles%202020%22"
})
check("query builder keeps URL hint first", queries[0] == "Leflaive Pucelles 2020", True)
check("query builder does not add LWIN as text query", all(q != "1234567" for q in queries), True)
queries_manual_style = _build_search_queries("Domaine Leflaive, Puligny-Montrachet, Pucelles", "2020", "")
check("manual-style exact full query is prioritized", queries_manual_style[0] == "Domaine Leflaive, Puligny-Montrachet, Pucelles 2020", True)
queries_two_part = _build_search_queries("Chateau La Fleur de Bouard, Lalande de Pomerol", "2020", "")
check("2-part region query builder prioritizes exact full query", queries_two_part[0] == "Chateau La Fleur de Bouard, Lalande de Pomerol 2020", True)
check("2-part region query builder adds producer-only vintage query", "Chateau La Fleur de Bouard 2020" in queries_two_part, True)
queries_with_hints = _build_search_queries(
    "Chateau La Fleur de Bouard, Lalande de Pomerol",
    "2020",
    "",
    {
        "jr_search_url": "https://www.jancisrobinson.com/tastings?search-full=%22Ch%20La%20Fleur%20de%20Bo%C3%BCard%202020%22",
        "jr_producer": "Ch La Fleur de Bouard",
        "jr_appellation": "Lalande-de-Pomerol",
    },
)
check("manual JR URL stays first", queries_with_hints[0] == "Ch La Fleur de Boüard 2020", True)
check("manual JR producer/appellation query is prioritized", "Ch La Fleur de Bouard, Lalande-de-Pomerol" in queries_with_hints[:3], True)
check("manual JR hints keep query count compact", len(queries_with_hints) <= 3, True)

print("\nâ•â• 2d2. XLSX manual JR hint parsing â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•")
wb = Workbook()
ws = wb.active
ws.append(["Publisher", "LWIN11", "Product_Name", "Vintage", "url on the jr", "wine on the jr", "Producer on jr", "appellation on jr ", "url on the rp"])
ws.append(["Jancis Robinson", "Optional if Name set", "Optional if LWIN set", "YYYY or NV", "", "", "", "", ""])
ws.append([
    "Jancis Robinson",
    "10098312020",
    "Chateau La Fleur de Bouard, Lalande de Pomerol 2020 (Magnum)",
    "2020",
    "https://www.jancisrobinson.com/tastings?search-full=%22Ch%20La%20Fleur%20de%20Bo%C3%BCard%202020%22",
    "",
    "Ch La Fleur de Boüard",
    "Lalande-de-Pomerol",
    "https://www.robertparker.com/wines/123456/test-slug",
])
buf = io.BytesIO()
wb.save(buf)
parsed_rows = xlsx_mod.parse_xlsx(buf.getvalue())
check("xlsx parser keeps one data row", len(parsed_rows) == 1, True)
check("xlsx parser strips trailing pack/year from name", parsed_rows[0].get("name") == "Chateau La Fleur de Bouard, Lalande de Pomerol", True)
check("xlsx parser stores JR producer hint", parsed_rows[0].get("search_hints", {}).get("jr_producer") == "Ch La Fleur de Boüard", True)
check("xlsx parser stores JR appellation hint", parsed_rows[0].get("search_hints", {}).get("jr_appellation") == "Lalande-de-Pomerol", True)
check("xlsx parser stores RP URL hint", parsed_rows[0].get("search_hints", {}).get("rp_search_url") == "https://www.robertparker.com/wines/123456/test-slug", True)

print("\n== 2d3. XLSX manual Decanter hint parsing ==")
wb_dc = Workbook()
ws_dc = wb_dc.active
ws_dc.append([
    "Publisher", "LWIN11", "Product_Name", "Vintage",
    "Name on Decanter", "URL on Decanter search page", "URL on Decanter sperate wine page"
])
ws_dc.append(["Decanter", "Optional if Name set", "Optional if LWIN set", "YYYY or NV", "", "", ""])
ws_dc.append([
    "Decanter",
    "10159552010",
    "Chateau Trotanoy, Pomerol 2010",
    "2010",
    "Château Trotanoy, Espérance de Trotanoy, Pomerol, 2010",
    "https://www.decanter.com/wine-reviews/search/term/chateau-trotanoy%2C-pomerol-2010/page/1/",
    "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-trotanoy-esperance-de-trotanoy-pomerol-2010-30840/",
])
buf_dc = io.BytesIO()
wb_dc.save(buf_dc)
parsed_dc_rows = xlsx_mod.parse_xlsx(buf_dc.getvalue())
check("xlsx parser keeps one Decanter data row", len(parsed_dc_rows) == 1, True)
check("xlsx parser keeps Decanter raw product name", parsed_dc_rows[0].get("raw_name") == "Chateau Trotanoy, Pomerol 2010", True)
check("xlsx parser stores Decanter wine-name hint", parsed_dc_rows[0].get("search_hints", {}).get("decanter_wine_name") == "Château Trotanoy, Espérance de Trotanoy, Pomerol, 2010", True)
check("xlsx parser stores Decanter search URL hint", "wine-reviews/search/term/" in (parsed_dc_rows[0].get("search_hints", {}).get("decanter_search_url") or ""), True)
check("xlsx parser stores Decanter detail URL hint", "chateau-trotanoy-esperance-de-trotanoy" in (parsed_dc_rows[0].get("search_hints", {}).get("decanter_review_url") or ""), True)
check(
    "Decanter ignores mismatched hinted wine name",
    _dc_preferred_query_name(
        "Pierre-Yves Colin-Morey, Chassagne-Montrachet 1er Cru, La Grande Montagne",
        2015,
        {"decanter_wine_name": "Bâtard-Montrachet Grand Cru 2015"},
    ) == "Pierre-Yves Colin-Morey, Chassagne-Montrachet 1er Cru, La Grande Montagne",
    True,
)
check(
    "Decanter ignores mismatched hinted direct URL",
    len(_dc_direct_detail_urls(
        "Pierre-Yves Colin-Morey, Chassagne-Montrachet 1er Cru, La Grande Montagne",
        2015,
        {"decanter_review_url": "https://www.decanter.com/wine-reviews/france/burgundy/batard-montrachet-grand-cru-2015-99999/"},
    )) == 0,
    True,
)
check(
    "Decanter ignores generic hinted direct URL when query vintage is specific",
    len(_dc_direct_detail_urls(
        "Le Clarence de Haut-Brion",
        2018,
        {"decanter_review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/le-clarence-de-haut-brion-99999/"},
    )) == 0,
    True,
)
check(
    "Decanter keeps explicit direct URL typed in the query itself",
    len(_dc_direct_detail_urls(
        "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-trotanoy-esperance-de-trotanoy-pomerol-2010-30840/",
        2010,
        None,
    )) == 1,
    True,
)
check(
    "Decanter ignores mismatched hinted search URL",
    len(_dc_manual_search_urls(
        "Domaine de la Romanée-Conti, La Tâche Grand Cru",
        2004,
        {"decanter_search_url": "https://www.decanter.com/wine-reviews/search/term/richebourg-grand-cru-2004/page/1/"},
    )) == 0,
    True,
)

print("\n== 2d4. Decanter query cleanup ==")
dc_queries = _build_dc_queries("Chateau Figeac Premier Grand Cru Classe A, Saint-Emilion Grand Cru", 2006)
check("Decanter keeps cleaned producer/appellation query", "Chateau Figeac, Saint-Emilion" in dc_queries, True)
check("Decanter adds cleaned vintage query", "Chateau Figeac, Saint-Emilion 2006" in dc_queries, True)
dc_queries_2 = _build_dc_queries("Chateau Latour-Martillac, Rouge Cru Classe, Pessac-Leognan", 2016)
check("Decanter drops descriptor-only middle segment", "Chateau Latour-Martillac, Pessac-Leognan" in dc_queries_2, True)
dc_queries_3 = _build_dc_queries("Chateau Malescot St. Exupery 3eme Cru Classe, Margaux", 2023)
check("Decanter expands saint abbreviation variant", any("Saint Exupery" in q for q in dc_queries_3), True)
dc_queries_4 = _build_dc_queries("Blason d'Issan, Margaux 2018 (Magnum)", 2018)
check("Decanter strips bottle size from query variants", all("Magnum" not in q for q in dc_queries_4), True)
check("Decanter avoids duplicate vintage after bottle-size cleanup", "Blason d'Issan, Margaux 2018 2018" not in dc_queries_4, True)
dc_html = """
<html><body>
  <h1>Château Trotanoy, Espérance de Trotanoy, Pomerol, 2010</h1>
  <div class="WineInfo_wineInfo__item__">
    <div class="type__">Producer</div><div class="value__">Château Trotanoy</div>
  </div>
  <div class="WineInfo_wineInfo__item__">
    <div class="type__">Appellation</div><div class="value__">Pomerol</div>
  </div>
  <script id="__NEXT_DATA__" type="application/json">
    {"props":{"pageProps":{"wine":{"vintage":2020,"primary_tasting":{"rounded_score":90,"consolidated_review":"Structured and firm.","drink_from":2019,"drink_to":2030,"published_at":"2019-06-12T15:25:08+01:00","scores":[{"review":"Structured and firm.","score":90,"judge":{"name":"Jane Anson"}}],"tasting":{"start_date":"2019-06-08T00:00:00+01:00"}}}}}}
  </script>
</body></html>
"""
dc_parsed = _dc_parse_detail_page(dc_html, "https://www.decanter.com/wine-reviews/example")
check("Decanter detail parser falls back to __NEXT_DATA__ score", dc_parsed.get("score_native") == 90, True)
check("Decanter detail parser falls back to __NEXT_DATA__ note", dc_parsed.get("note") == "Structured and firm.", True)
check("Decanter detail parser falls back to __NEXT_DATA__ reviewer", dc_parsed.get("reviewer") == "Jane Anson", True)
check("Decanter detail parser falls back to __NEXT_DATA__ vintage", dc_parsed.get("vintage_src") == 2020, True)

print("\n== 2d5. Decanter vintage guard ==")
generic_dc_review = {
    "wine_name_src": "Château Lynch-Moussas, Pauillac, Bordeaux, France",
    "producer": "Château Lynch-Moussas",
    "appellation": "Pauillac",
    "region": "Bordeaux",
    "country": "France",
    "vintage_src": None,
    "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-lynch-moussas-bordeaux-france-1009/",
}
check(
    "Decanter rejects generic no-vintage page for vintage query",
    _dc_matches_query("Chateau Lynch-Moussas 5eme Cru Classe, Pauillac", 2021, generic_dc_review),
    False,
)
check(
    "Decanter still trusts explicit direct URLs when vintage is missing on page",
    _dc_matches_query(
        "Chateau Lynch-Moussas 5eme Cru Classe, Pauillac",
        2021,
        generic_dc_review,
        direct_url="https://www.decanter.com/wine-reviews/france/bordeaux/chateau-lynch-moussas-bordeaux-france-1009/",
        allow_missing_vintage=True,
    ),
    True,
)
check(
    "Decanter direct URLs still reject a different producer",
    _dc_matches_query(
        "Chateau Dassault, Saint-Emilion Grand Cru",
        2020,
        {
            "wine_name_src": "Chateau Fombrauge, St-Emilion Grand Cru 2020",
            "producer": "Chateau Fombrauge",
            "appellation": "St-Emilion",
            "vintage_src": 2020,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-fombrauge-st-emilion-grand-cru-2020-99999/",
        },
        direct_url="https://www.decanter.com/wine-reviews/france/bordeaux/chateau-fombrauge-st-emilion-grand-cru-2020-99999/",
        allow_missing_vintage=True,
    ),
    False,
)
check(
    "Decanter direct URLs reject same producer when appellation drifts",
    _dc_matches_query(
        "Chateau Margaux, Margaux",
        2020,
        {
            "wine_name_src": "Chateau Margaux, St-Emilion Grand Cru 2020",
            "producer": "Chateau Margaux",
            "appellation": "St-Emilion",
            "vintage_src": 2020,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-margaux-st-emilion-grand-cru-2020-99998/",
        },
        direct_url="https://www.decanter.com/wine-reviews/france/bordeaux/chateau-margaux-st-emilion-grand-cru-2020-99998/",
        allow_missing_vintage=True,
    ),
    False,
)
check(
    "Decanter direct URLs reject rouge-vs-blanc mismatches",
    _dc_matches_query(
        "Domaine de Chevalier, Rouge, Pessac-Leognan",
        2020,
        {
            "wine_name_src": "Domaine de Chevalier, Blanc, Pessac-Leognan 2020",
            "producer": "Domaine de Chevalier",
            "appellation": "Pessac-Leognan",
            "colour": "white",
            "vintage_src": 2020,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/domaine-de-chevalier-blanc-pessac-leognan-2020-99997/",
        },
        direct_url="https://www.decanter.com/wine-reviews/france/bordeaux/domaine-de-chevalier-blanc-pessac-leognan-2020-99997/",
        allow_missing_vintage=True,
    ),
    False,
)

print("\n== 2d6. Decanter candidate ordering ==")
ordered = _dc_sort_candidates([
    {"title": "Chateau Lafaurie-Peyraguey, Sauternes, 1er Cru Classe, 2009", "url": "https://example.test/2009", "vintage": 2009, "rank_score": 99},
    {"title": "Chateau Lafaurie-Peyraguey, Sauternes, 1er Cru Classe, 2016", "url": "https://example.test/2016", "vintage": 2016, "rank_score": 70},
], 2016)
check("Decanter prefers exact-vintage candidate over higher-score wrong vintage", ordered[0].get("vintage") == 2016, True)
pool = _dc_candidate_review_pool([
    {"title": f"Wine {i}", "url": f"https://example.test/{i}", "vintage": 2010, "rank_score": 100 - i}
    for i in range(12)
], 2010)
check("Decanter review pool checks deeper than the old top-8 limit", len(pool) == 12, True)
check(
    "Decanter accepts detail-page facts when title omits appellation",
    _dc_matches_query(
        "Chateau Trotanoy, Pomerol",
        2010,
        {
            "wine_name_src": "Château Trotanoy, Espérance de Trotanoy, Bordeaux, France, 2010",
            "producer": "Château Trotanoy",
            "appellation": "Pomerol",
            "region": "Bordeaux",
            "country": "France",
            "vintage_src": 2010,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-trotanoy-esperance-de-trotanoy-pomerol-2010-30840/",
        },
    ),
    True,
)
check(
    "Decanter treats accented producer and extra subtitle words as same identity",
    _dc_matches_query(
        "Chateau Trotanoy, Pomerol",
        2010,
        {
            "wine_name_src": "Château Trotanoy, Espérance de Trotanoy, Bordeaux, France, 2010",
            "vintage_src": 2010,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-trotanoy-esperance-de-trotanoy-pomerol-2010-30840/",
            "producer": "Château Trotanoy",
            "appellation": "Pomerol",
        },
    ),
    True,
)
check(
    "Decanter rejects named-estate neighbour with extra producer token",
    _dc_matches_query(
        "Chateau Grand Corbin, Saint-Emilion Grand Cru",
        2018,
        {
            "wine_name_src": "Château Grand Corbin Despagne, St-Émilion, Grand Cru Classé, 2018",
            "producer": "Château Grand Corbin Despagne",
            "appellation": "St-Émilion",
            "vintage_src": 2018,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-grand-corbin-despagne-st-emilion-grand-cru-classe-44780/",
        },
    ),
    False,
)
check(
    "Decanter rejects red-vs-white confusion when colour sits inside descriptor segment",
    _dc_matches_query(
        "Domaine de Chevalier, Rouge Cru Classe, Pessac-Leognan",
        2015,
        {
            "wine_name_src": "Domaine de Chevalier, Blanc, Pessac-Léognan Cru Classé de Graves, Bordeaux, France 2015",
            "producer": "Domaine de Chevalier",
            "appellation": "Pessac-Léognan",
            "colour": "white",
            "vintage_src": 2015,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/domaine-de-chevalier-blanc-pessac-leognan-cru-classe-de-102661/",
        },
    ),
    False,
)
check(
    "Decanter rejects parent wine when query is a second label",
    _dc_matches_query(
        "La Reserve de Leoville Barton, Saint-Julien",
        2020,
        {
            "wine_name_src": "Château Léoville Barton, St-Julien, 2ème Cru Classé, Bordeaux, France 2020",
            "producer": "Château Léoville Barton",
            "appellation": "St-Julien",
            "vintage_src": 2020,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-leoville-barton-st-julien-2eme-cru-classe-2020-48335/",
        },
    ),
    False,
)
check(
    "Decanter rejects second wine when title carries hidden sub-label",
    _dc_matches_query(
        "Chateau Figeac Premier Grand Cru Classe A, Saint-Emilion Grand Cru",
        2018,
        {
            "wine_name_src": "Château Figeac, Petit Figeac, St-Émilion, Grand Cru, Bordeaux, France 2018",
            "producer": "Château Figeac",
            "appellation": "St-Émilion",
            "region": "Bordeaux",
            "country": "France",
            "vintage_src": 2018,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-figeac-petit-figeac-st-emilion-grand-cru-2018-71139/",
        },
    ),
    False,
)
check(
    "Decanter rejects classification letter mismatches",
    _dc_matches_query(
        "Chateau Figeac Premier Grand Cru Classe A, Saint-Emilion Grand Cru",
        2018,
        {
            "wine_name_src": "Château Figeac, St-Émilion, 1er Grand Cru Classé B 2018",
            "producer": "Château Figeac",
            "appellation": "St-Émilion",
            "vintage_src": 2018,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/chateau-figeac-st-emilion-1er-grand-cru-classe-b-2018-29430/",
        },
    ),
    False,
)
check(
    "Decanter ignores bottle-size suffixes during final match checks",
    _dc_matches_query(
        "Petrus, Pomerol 2008 (Magnum)",
        2008,
        {
            "wine_name_src": "Petrus, Pomerol, Bordeaux, France, 2008",
            "producer": "Petrus",
            "appellation": "Pomerol",
            "vintage_src": 2008,
            "review_url": "https://www.decanter.com/wine-reviews/france/bordeaux/petrus-pomerol-bordeaux-france-2008-18360/",
        },
    ),
    True,
)
check(
    "Decanter rejects producer-overlap hit when appellation drifts",
    _dc_matches_query(
        "Chateau Moulin a Vent, Moulis en Medoc",
        2016,
        {
            "wine_name_src": "Chateau du Moulin-a-Vent, Moulin-a-Vent, Beaujolais 2016",
            "producer": "Chateau du Moulin-a-Vent",
            "appellation": "Moulin-a-Vent",
            "region": "Beaujolais",
            "country": "France",
            "vintage_src": 2016,
            "review_url": "https://www.decanter.com/wine-reviews/france/burgundy/chateau-du-moulin-a-vent-beaujolais-moulin-a-vent-2016-62348/",
        },
    ),
    False,
)
old_openai_key = os.environ.get("OPENAI_API_KEY")
old_dc_openai = os.environ.get("DECANTER_OPENAI_FALLBACK")
try:
    os.environ["OPENAI_API_KEY"] = "test-key"
    os.environ.pop("DECANTER_OPENAI_FALLBACK", None)
    check("Decanter OpenAI fallback auto-enables when key exists", _dc_openai_fallback_enabled(), True)
    os.environ["DECANTER_OPENAI_FALLBACK"] = "0"
    check("Decanter OpenAI fallback still honours explicit disable flag", _dc_openai_fallback_enabled(), False)
finally:
    if old_openai_key is None:
        os.environ.pop("OPENAI_API_KEY", None)
    else:
        os.environ["OPENAI_API_KEY"] = old_openai_key
    if old_dc_openai is None:
        os.environ.pop("DECANTER_OPENAI_FALLBACK", None)
    else:
        os.environ["DECANTER_OPENAI_FALLBACK"] = old_dc_openai

print("\n══ 2e. JR plausible-match fallback (_jr_plausible_match) ════════════════")
check(
    "JR allows strong abbreviated producer/title matches",
    _jr_plausible_match(
        "Azienda Agricola Salvioni La Cerbaiola, Brunello di Montalcino",
        "Salvioni Brunello",
        {"vintage": ["2020"]},
        "2020",
        "",
    ),
    True,
)
check(
    "JR still rejects unrelated weak matches",
    _jr_plausible_match(
        "Azienda Agricola Salvioni La Cerbaiola, Brunello di Montalcino",
        "Dry Amber Reserve",
        {"vintage": ["2020"]},
        "2020",
        "",
    ),
    False,
)

print("\n══ 2f. JR strict structured match ═══════════════════════════════════════")
strict_candidate = {
    "producer": "Domaine Leflaive",
    "wine_name": "Puligny-Montrachet",
    "appellation": "Pucelles",
    "vintage": "2020",
    "match_text": "Domaine Leflaive Puligny-Montrachet Pucelles",
}
check("strict structured match accepts exact identity", _strict_match_structured(structured, strict_candidate), True)

print("\nâ•â• 2f2. JR benchmark candidate ranking â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•")
benchmark_query = _parse_query_structured("Chateau La Fleur de Bouard, Lalande de Pomerol", "2020")
benchmark_good = {
    "producer": "Ch La Fleur de Bouard",
    "wine_name": "",
    "title": "",
    "appellation": "Lalande-de-Pomerol",
    "vintage": "2020",
    "match_text": "Ch La Fleur de Bouard Lalande-de-Pomerol",
    "rank": 70.0,
    "date_tasted": "08 Jun 2023",
}
benchmark_bad = {
    "producer": "",
    "wine_name": "",
    "title": "",
    "appellation": "Lalande-de-Pomerol",
    "vintage": "2020",
    "match_text": "Random Lalande-de-Pomerol",
    "rank": 70.0,
    "date_tasted": "08 Jun 2023",
}
check(
    "benchmark scorer prefers same producer over region-only match",
    _candidate_prerank_score(benchmark_query, benchmark_good) > _candidate_prerank_score(benchmark_query, benchmark_bad),
    True,
)
check(
    "benchmark gate rejects region-only false positive",
    _jr_candidate_passes([benchmark_query], benchmark_bad),
    False,
)
check(
    "benchmark gate accepts producer/appellation match",
    _jr_candidate_passes([benchmark_query], benchmark_good),
    True,
)

print("\n══ 2g. JR candidate selection (_jr_select_candidate) ════════════════════")
_orig_pick = jr_mod._jr_llm_pick_candidate
_orig_verify = jr_mod._jr_llm_verify_exact_match
_orig_strict = jr_mod.JR_STRICT_MODE
try:
    c1 = {"match_text": "Producer A Wine A", "rank": 80.0}
    c2 = {"match_text": "Producer A Wine B", "rank": 76.0}
    jr_mod.JR_STRICT_MODE = True
    jr_mod._jr_llm_pick_candidate = lambda *args, **kwargs: c2
    jr_mod._jr_llm_verify_exact_match = lambda *args, **kwargs: True
    picked = jr_mod._jr_select_candidate("Producer A, Region, Wine B", "2020", [c1, c2])
    check("LLM can resolve ambiguous JR top results", picked is c2, True)
finally:
    jr_mod._jr_llm_pick_candidate = _orig_pick
    jr_mod._jr_llm_verify_exact_match = _orig_verify
    jr_mod.JR_STRICT_MODE = _orig_strict

print("\n══ 2h. JR search aggregation (jr_msearch) ═══════════════════════════════")
_orig_do = jr_mod._do_msearch
try:
    calls = []
    hit_a = {"url": ["/a"], "title": ["A"], "producer": ["Prod"], "wine_name": ["Wine A"], "appellation": ["App"], "lwin": ["1111111"], "vintage": ["2020"]}

    def fake_do(_session, payload):
        calls.append(payload)
        return [hit_a] if len(calls) == 1 else []

    jr_mod._do_msearch = fake_do
    hits = jr_mod.jr_msearch(object(), "Producer, Region, Wine", "2020", "")
    check("JR keeps first relevant compact ES hit set", len(hits) == 1, True)
finally:
    jr_mod._do_msearch = _orig_do

print("\n══ 2i. JR search_wine prefers structured ES search ═══════════════════════")
_orig_jr_msearch = jr_mod.jr_msearch
_orig_es_candidate = jr_mod._jr_es_candidate
_orig_review_from_candidate = jr_mod._jr_review_from_es_candidate
try:
    jr_mod.jr_msearch = lambda *args, **kwargs: [{"title": ["Used"]}]
    jr_mod._jr_es_candidate = lambda *args, **kwargs: {"rank": 77.0, "review_url": "https://example.test/review", "src": {}}
    jr_mod._jr_review_from_es_candidate = lambda *args, **kwargs: {"score_20": 16, "wine_name_jr": "ES Pick", "review_url": "https://example.test/review"}
    hits = jr_mod.search_wine(object(), "Producer, Region, Wine", "2020", "")
    check("JR uses ES result", len(hits) == 1 and hits[0].get("wine_name_jr") == "ES Pick", True)
finally:
    jr_mod.jr_msearch = _orig_jr_msearch
    jr_mod._jr_es_candidate = _orig_es_candidate
    jr_mod._jr_review_from_es_candidate = _orig_review_from_candidate

print("\n══ 2j. JR search_wine falls back to manual search ════════════════════════")
_orig_jr_msearch = jr_mod.jr_msearch
try:
    jr_mod.jr_msearch = lambda *args, **kwargs: []
    hits = jr_mod.search_wine(object(), "Producer, Region, Wine", "2020", "")
    check("JR returns empty when ES has no result", len(hits) == 0, True)
finally:
    jr_mod.jr_msearch = _orig_jr_msearch


print("\n══ 2k. XLSX runner passes JR search hints ═════════════════")
_orig_get_session = None
_orig_jr_search = None
_orig_upsert_review_wine = None
try:
    import services.session_service as session_service
    import models.wine_model as wine_model
    _orig_get_session = session_service.get_session
    _orig_jr_search = jr_mod.search_wine
    _orig_upsert_review_wine = wine_model.upsert_review_wine
    captured = {}
    saved_reviews = []

    def fake_get_session(_source_key):
        return object()

    def fake_search(session, name, vintage, lwin="", search_hints=None):
        captured["search_hints"] = dict(search_hints or {})
        return [{"score_20": 16.0, "date_tasted": "08 Jun 2023", "review_url": "https://example.test/review"}]

    def fake_upsert_review_wine(lwin_full, wine_data, review_data, upload_batch=""):
        saved_reviews.append({
            "lwin": lwin_full,
            "wine_data": dict(wine_data or {}),
            "review_data": dict(review_data or {}),
            "upload_batch": upload_batch,
        })
        return {"action": "updated", "wine_id": 1}

    session_service.get_session = fake_get_session
    jr_mod.search_wine = fake_search
    wine_model.upsert_review_wine = fake_upsert_review_wine

    job_id = xlsx_mod.create_job(
        buf.getvalue(),
        [{
            "row_idx": 3,
            "name": "Chateau La Fleur de Bouard, Lalande de Pomerol",
            "vintage": "2020",
            "lwin": "10098312020",
            "search_hints": {
                "jr_search_url": "https://www.jancisrobinson.com/tastings?search-full=%22Ch%20La%20Fleur%20de%20Bo%C3%BCard%202020%22",
                "jr_producer": "Ch La Fleur de Bo\u00fcard",
                "jr_appellation": "Lalande-de-Pomerol",
                "rp_search_url": "https://www.robertparker.com/wines/123456/test-slug",
            },
        }],
    )
    xlsx_mod.run_job(job_id, source_key="jancisrobinson", sleep_sec=0)
    check("xlsx runner forwards JR producer hint", captured.get("search_hints", {}).get("jr_producer") == "Ch La Fleur de Bo\u00fcard", True)
    check("xlsx runner forwards JR URL hint", "search-full" in (captured.get("search_hints", {}).get("jr_search_url") or ""), True)
    check("xlsx runner forwards RP URL hint", captured.get("search_hints", {}).get("rp_search_url") == "https://www.robertparker.com/wines/123456/test-slug", True)
    check("xlsx runner saves review URL into DB payload", saved_reviews[0]["review_data"].get("review_url") == "https://example.test/review", True)
finally:
    if _orig_get_session is not None:
        session_service.get_session = _orig_get_session
    if _orig_jr_search is not None:
        jr_mod.search_wine = _orig_jr_search
    if _orig_upsert_review_wine is not None:
        wine_model.upsert_review_wine = _orig_upsert_review_wine

print("\n== 2k2. XLSX runner strips paywalls, stubs, and duplicate session notes ==")
_orig_get_session = None
_orig_jr_search = None
_orig_upsert_review_wine = None
try:
    import services.session_service as session_service
    import models.wine_model as wine_model
    _orig_get_session = session_service.get_session
    _orig_jr_search = jr_mod.search_wine
    _orig_upsert_review_wine = wine_model.upsert_review_wine
    saved_reviews = []

    def fake_get_session(_source_key):
        return object()

    duplicate_note = (
        "60% Sauvignon Blanc, 40% Semillon. Lovely aroma of zesty lemon cream, "
        "with grapefruit, cedar and vanilla. Fresh and long on the finish."
    )

    def fake_search(session, name, vintage, lwin="", search_hints=None):
        if "Wine A" in name:
            return [{
                "score_20": 16.5,
                "date_tasted": "22 Apr 2015",
                "reviewer": "Julia Harding MW",
                "tasting_note": duplicate_note,
                "review_url": "https://example.test/a",
            }]
        if "Wine B" in name:
            return [{
                "score_20": 16.5,
                "date_tasted": "22 Apr 2015",
                "reviewer": "Julia Harding MW",
                "tasting_note": duplicate_note,
                "review_url": "https://example.test/b",
            }]
        if "Wine C" in name:
            return [{
                "score_20": 17.0,
                "date_tasted": "05 Jan 2026",
                "reviewer": "Jancis Robinson",
                "tasting_note": "Become a member to read more",
                "review_url": "https://example.test/paywall",
            }]
        return [{
            "score_20": 17.0,
            "date_tasted": "10 Dec 2007",
            "reviewer": "Jancis Robinson",
            "tasting_note": "Magnum",
            "review_url": "https://example.test/stub",
        }]

    def fake_upsert_review_wine(lwin_full, wine_data, review_data, upload_batch=""):
        saved_reviews.append(dict(review_data or {}))
        return {"action": "updated", "wine_id": 1}

    session_service.get_session = fake_get_session
    jr_mod.search_wine = fake_search
    wine_model.upsert_review_wine = fake_upsert_review_wine

    workbook = Workbook()
    ws2 = workbook.active
    ws2.append(["Publisher", "LWIN11", "Product_Name", "Vintage", "Critic_Name", "Score", "Review_Date", "Review", "Source_URL"])
    ws2.append(["Jancis Robinson", "Optional if Name set", "Optional if LWIN set", "YYYY or NV", "", "", "", "", ""])
    ws2.append(["Jancis Robinson", "11111112020", "Wine A 2020", "2020", "", "", "", "", ""])
    ws2.append(["Jancis Robinson", "22222222021", "Wine B 2021", "2021", "", "", "", "", ""])
    ws2.append(["Jancis Robinson", "33333332022", "Wine C 2022", "2022", "", "", "", "", ""])
    ws2.append(["Jancis Robinson", "44444441996", "Wine D 1996", "1996", "", "", "", "", ""])
    buf2 = io.BytesIO()
    workbook.save(buf2)
    wines2 = xlsx_mod.parse_xlsx(buf2.getvalue())
    job_id = xlsx_mod.create_job(buf2.getvalue(), wines2)
    xlsx_mod.run_job(job_id, source_key="jancisrobinson", sleep_sec=0)
    internal_results = xlsx_mod._jobs[job_id]["results"]

    check("xlsx runner blanks both duplicate session notes", all(not internal_results[i].get("note") for i in (0, 1)), True)
    check("xlsx runner blanks paywall placeholder note", internal_results[2].get("note") == "", True)
    check("xlsx runner blanks short stub note", internal_results[3].get("note") == "", True)
    check("xlsx runner still preserves review URL after note cleanup", saved_reviews[2].get("review_url") == "https://example.test/paywall", True)
finally:
    if _orig_get_session is not None:
        session_service.get_session = _orig_get_session
    if _orig_jr_search is not None:
        jr_mod.search_wine = _orig_jr_search
    if _orig_upsert_review_wine is not None:
        wine_model.upsert_review_wine = _orig_upsert_review_wine

# ─── 3. RP query builder ──────────────────────────────────────────────────────
print("\n══ 3. RP query builder (_build_rp_queries) ══════════════════════════")

WINES = [
    ("Charles Noellat, Nuits-Saint-Georges, Blanc", 2006),
    ("Château Margaux",                              2015),
    ("Giacomo Conterno, Barolo, Monfortino Riserva", 2016),
    ("Domaine Leroy, Chambolle-Musigny, Les Charmes",2019),
    ("Pétrus",                                       2012),
    ("Penfolds Grange",                              2018),
]

for name, vintage in WINES:
    qs = _build_rp_queries(name, vintage)
    print(f"\n  Wine   : {name} ({vintage})")
    for i, q in enumerate(qs):
        print(f"  Query {i+1}: {q!r}")

print("\nâ•â• Optional manual benchmark workbook â•â•")
manual_workbook = r"C:\Users\RTS84\Downloads\ruepinard-attachments (1)\Jancis Robinson for the research the data 1-200.xlsx"
if os.path.exists(manual_workbook):
    with open(manual_workbook, "rb") as fh:
        benchmark_rows = xlsx_mod.parse_xlsx(fh.read())
    check("manual benchmark workbook keeps 198 wine rows", len(benchmark_rows) == 198, True)
    first_row_queries = _build_search_queries(benchmark_rows[0]["name"], benchmark_rows[0]["vintage"], benchmark_rows[0]["lwin"])
    check("manual benchmark first row starts with exact full search", first_row_queries[0] == "Chateau La Fleur de Bouard, Lalande de Pomerol 2020", True)
    check("manual benchmark query list stays compact", len(first_row_queries) <= 12, True)
else:
    print("  - skipped (manual benchmark workbook not present)")


# ─── 4. Live RP search ────────────────────────────────────────────────────────
if "live" in sys.argv:
    print("\n══ 4. Live RP search ════════════════════════════════════════════════")
    try:
        from sources.robertparker import load_session, search_wine as rp_search

        session_rp = load_session("cookies/robertparker.json")
        tests = [
            ("Charles Noellat, Nuits-Saint-Georges, Blanc", 2006, None),
            ("Château Margaux",                              2015, None),
            ("Giacomo Conterno, Barolo, Monfortino Riserva", 2016, None),
        ]
        for name, vintage, lwin in tests:
            print(f"\n  Searching: {name} {vintage}")
            results = rp_search(session_rp, name, vintage, lwin, sleep_sec=1.5)
            if results:
                r = results[0]
                print(f"  {PASS}  Found: {r.get('wine_name_src')} | "
                      f"score={r.get('score_native')}/100 | "
                      f"reviewer={r.get('reviewer')}")
            else:
                print(f"  {FAIL}  Not found")
    except Exception as e:
        print(f"  {FAIL}  {type(e).__name__}: {e}")


# ─── 5. Live JR search ────────────────────────────────────────────────────────
if "live" in sys.argv and "jr" in sys.argv:
    print("\n══ 5. Live JR search ════════════════════════════════════════════════")
    try:
        from maaike_phase1 import load_session as jr_load, search_wine as jr_search

        session_jr = jr_load("real_cookies.json")
        tests = [
            ("Charles Noellat, Nuits-Saint-Georges, Blanc", "2006", "LWIN285595020060600750"),
            ("Château Margaux",                              "2015", ""),
        ]
        for name, vintage, lwin in tests:
            print(f"\n  Searching: {name} {vintage}")
            results = jr_search(session_jr, name, vintage, lwin)
            if results:
                r = results[0]
                print(f"  {PASS}  Found: {r.get('wine_name_jr')} | "
                      f"score={r.get('score_20')}/20 | "
                      f"reviewer={r.get('reviewer')}")
            else:
                print(f"  {FAIL}  Not found")
    except Exception as e:
        print(f"  {FAIL}  {type(e).__name__}: {e}")


print("\n")
